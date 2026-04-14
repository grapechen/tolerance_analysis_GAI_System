import sys, io
import os
import json
import copy
import pandas as pd
from flask import Flask, request, jsonify, render_template, Response
from flask_cors import CORS
import ollama
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# 假設這些是你專案中自定義的模組，請確保它們與 app.py 在可被解析的路徑中
from scripts.triplets_extractor import get_mating_constraints
import graph_rag
from rag_engine import ask_rag_engine
from scripts.dsl_builder import build_full_dsl

# STEP PMI 服務 (Phase 1 整合)
from step_service import (
    route_upload_step,
    route_parse_pmi,
    route_get_geometry,
    route_get_pmi_list,
    route_highlight_pmi,
    route_run_asm_worker,
    route_get_6dof,
    route_export_step_csv,
    route_progress_sse,
    route_progress_status
)

# 處理 Windows 終端機編碼問題 
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# 載入環境變數 
load_dotenv()

app = Flask(__name__)
CORS(app)  # 允許來自前端的跨網域請求 (CORS) 

def get_available_models():
    """
    取得目前本機與已設定的可用 AI 模型列表。
    """
    try:
        client = ollama.Client(host="http://localhost:11434")
        models_info = client.list()
        model_names = []
        for m in models_info.models:
            m_name = None
            if hasattr(m, "model"):
                m_name = m.model
            elif hasattr(m, "name"):
                m_name = m.name
            elif isinstance(m, dict):
                m_name = m.get("model") or m.get("name") 
            
            if m_name:
                model_names.append(m_name) 

        # 定義雲端模型前綴 
        cloud_model_prefixes = [
            "gpt-oss", "qwen3-vl", "qwen3-v1", "ministral-3", "qwen3-coder",
            "glm-5", "glm-4.7", "glm-4.6", "glm-4", "deepseek-v3.2",
            "deepseek-v3.1", "deepseek3.1", "deepseek-v3", "minimax-m2",
            "minimax", "gemini-3", "kimi", "qwen3.5", "nemotron-3",
        ]

        def is_cloud_model(name):
            name_lower = name.lower()
            if "-cloud" in name_lower or ":cloud" in name_lower:
                return True 
            for prefix in cloud_model_prefixes:
                if name_lower.startswith(prefix):
                    return True 
            return False 

        def model_sort_key(name):
            if is_cloud_model(name):
                return (0, name.lower()) 
            else:
                return (1, name.lower()) 

        manual_cloud_models = [
            "gpt-oss:120b-cloud", "deepseek3.1:671b-cloud", "qwen3-coder:480b-cloud",
            "ministral-3:8b-cloud", "glm-4.7:cloud", "minimax-m2:cloud",
        ] 

        import re
        final_model_dict = {}
        for m in model_names + manual_cloud_models:
            m_lower = m.lower()
            if "gemini" in m_lower:
                continue 
            match = re.match(r"^([a-z\-]+)(?:[\d\.\-v]*)(?:[:\-].*)?$", m_lower) 
            if match:
                base_family = match.group(1).strip("-")
                if base_family.startswith("deepseek"): base_family = "deepseek"
                elif base_family.startswith("qwen"): base_family = "qwen"
                elif base_family.startswith("glm"): base_family = "glm"
                elif base_family.startswith("gpt"): base_family = "gpt" 
            else:
                base_family = m_lower.split(":")[0] 

            # [核心修正] 本地模型不進行 Family 去重，確保所有具體量化版 (Tags) 都能顯示
            # 雲端模型則維持 family 唯一性
            model_key = base_family if "cloud" in m_lower else m
            
            if model_key not in final_model_dict:
                final_model_dict[model_key] = m 
            else:
                current_best = final_model_dict[model_key]
                is_m_cloud = "cloud" in m_lower 
                is_curr_cloud = "cloud" in current_best.lower() 
                if is_m_cloud and not is_curr_cloud:
                    final_model_dict[model_key] = m 
                elif is_m_cloud == is_curr_cloud:
                    if len(m) > len(current_best):
                        final_model_dict[model_key] = m 

        model_names = list(final_model_dict.values()) 
        model_names.sort(key=model_sort_key) 
    except Exception as e:
        print(f"Error fetching models: {e}") 
        model_names = ["llama3.1:8b"] 

    current_model = None
    preferred_cloud = [
        "gemma3:4b", "gemma3:12b", "minimax-m2:cloud",
        "gpt-oss:120b-cloud", "ministral-3:8b-cloud", "qwen3-coder:480b-cloud",
    ] 
    for preferred in preferred_cloud:
        if any(str(m) == str(preferred) for m in model_names):
            current_model = preferred 
            break 

    if not current_model:
        for m in model_names:
            if m.startswith("gemma3:") or m.startswith("llama3"):
                current_model = m 
                break 

    if not current_model:
        current_model = model_names[0] if model_names else "llama3.1:8b" 

    return model_names, current_model 


@app.route("/")
def home():
    """繁體中文版路由"""
    model_names, current_model = get_available_models() 
    constraints = get_mating_constraints() 
    # 改用 render_template 呼叫獨立的 HTML 檔案
    return render_template(
        "index.html", 
        models=model_names, 
        current_model=current_model, 
        lang="zh-TW",
        mating_constraints=constraints
    ) 


@app.route("/en")
def home_en():
    """英文版路由"""
    model_names, current_model = get_available_models() 
    constraints = get_mating_constraints() 
    return render_template(
        "index.html", 
        models=model_names, 
        current_model=current_model, 
        lang="en",
        mating_constraints=constraints
    ) 


@app.route("/api/chat", methods=["POST"])
def api_chat():
    data = request.get_json(force=True)
    user_msg = data.get("message", "")
    model_name = data.get("model", "llama3:8b-instruct-q4_K_M")
    history = data.get("history", [])
    lang = data.get("lang", "zh-TW")
    current_analysis = data.get("current_analysis", None)
    current_path = data.get("current_path", None)
    current_allocation = data.get("current_allocation", None)
    current_pmi_session_id = data.get("current_pmi_session_id", None)  # [Phase 4] PMI Session

    if not user_msg:
        reply_msg = "Please enter a message" if lang == "en" else "請輸入訊息" 
        return jsonify({"reply": reply_msg}), 400 

    model_lower = model_name.lower() 
    is_cloud = (
        "-cloud" in model_lower
        or ":cloud" in model_lower
        or model_lower.startswith("gpt-oss")
        or model_lower.startswith("qwen3-vl")
        or model_lower.startswith("qwen3-v1")
        or model_lower.startswith("ministral-3")
        or model_lower.startswith("qwen3-coder")
        or model_lower.startswith("glm-4")
        or model_lower.startswith("deepseek")
        or model_lower.startswith("minimax")
    ) 

    CLOUD_OLLAMA_URL = "http://localhost:11434" 
    base_url = CLOUD_OLLAMA_URL if is_cloud else "http://localhost:11434" 

    print(f"[INFO] 接收到對話請求 - 訊息: '{user_msg}', 模型: {model_name}, 網址: {base_url}") 

    try:
        reply, bom_intent = ask_rag_engine(
            user_msg, model_name=model_name, base_url=base_url, history=history, lang=lang,
            current_analysis=current_analysis, current_path=current_path,
            current_allocation=current_allocation, current_pmi_session=current_pmi_session_id  # [Phase 4]
        )
    except Exception as e:
        import sys
        with open("sys_exec.txt", "w", encoding="utf-8") as f:
            f.write(f"exe: {sys.executable}\npath: {sys.path}\nerror: {e}") 
        print(f"[WARN] GraphRAG 匯入或執行失敗: {e}") 
        print(f"[WARN] 正在使用的 Python: {sys.executable}") 
        reply = f"[ERROR] 圖資料庫 (GraphRAG) 執行發生錯誤: {e}。請聯絡系統管理員。" 
        bom_intent = {} 

    return jsonify({"reply": reply, "intent": bom_intent}) 


@app.route("/api/analyze_tolerance_stream", methods=["GET"])
def analyze_tolerance_stream():
    """SSE 端點：接收 editorPathData 執行 Jacobian + RSS + Monte Carlo，推播進度。"""
    raw = request.args.get("pathData", "[]")
    try:
        path_data = json.loads(raw)
    except Exception:
        return jsonify({"error": "pathData JSON 解析失敗"}), 400

    # 解析額外參數 (取樣數, sigma, 分布模式)
    try:
        mc_samples = int(request.args.get("n_samples", 10000))
        mc_sigma   = float(request.args.get("sigma", 3.0))
        mc_dist    = int(request.args.get("dist_type", 0)) # 0:均勻, 1:常態
    except ValueError:
        mc_samples, mc_sigma, mc_dist = 10000, 3.0, 0

    try:
        from analysis_service import analyze_stream
    except ImportError as e:
        return jsonify({"error": f"analysis_service 模組載入失敗: {e}"}), 500

    return Response(
        analyze_stream(path_data, mc_samples=mc_samples, mc_sigma=mc_sigma, mc_dist=mc_dist),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.route("/api/import_excel", methods=["POST"])
def import_excel():
    """接收上傳的 Excel 並解析為 JSON 路徑"""
    if "file" not in request.files:
        return jsonify({"error": "未提供檔案"}), 400
    
    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "未選擇檔案"}), 400
        
    try:
        from analysis_service import parse_excel_to_path
        path_data = parse_excel_to_path(file.read())
        return jsonify({"pathData": path_data})
    except Exception as e:
        return jsonify({"error": f"解析 Excel 失敗: {str(e)}"}), 500


@app.route("/api/machines", methods=["GET"])
def get_machines():
    file_path = os.path.join(os.path.dirname(__file__), "data", "machines_data.json") 
    if not os.path.exists(file_path):
        return jsonify({"ok": False, "msg": "找不到機台資料"}), 404 

    try:
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f) 
        return jsonify(data) 
    except Exception as e:
        return jsonify({"ok": False, "msg": f"解析資料失敗: {str(e)}"}), 500 


@app.route("/api/export_tolerance_csv", methods=["POST"])
def export_tolerance_csv():
    data = request.get_json() 
    path_data = data.get("pathData", []) 
    lang = data.get("lang", "zh-TW") 

    rows = [] 
    for item in path_data:
        if item.get("type") == "feature":
            rows.append({
                "路徑代碼" if lang != 'en' else "Path Code": item.get("name"), 
                "數值(平移、旋轉、公差值)" if lang != 'en' else "Value(tra/rot/tol)": item.get("val", 0.01), 
                "偏差值(公差帶偏移值)" if lang != 'en' else "Bias(offset)": item.get("bias", 0), 
                "角度公差轉換距離" if lang != 'en' else "Ang Tol Dist": item.get("dist", "") or "", 
            }) 
        elif item.get("type") == "spatial":
            rows.append({
                "路徑代碼" if lang != 'en' else "Path Code": item.get("axis"), 
                "數值(平移、旋轉、公差值)" if lang != 'en' else "Value(tra/rot/tol)": item.get("val", 0.0), 
                "偏差值(公差帶偏移值)" if lang != 'en' else "Bias(offset)": item.get("bias", 0), 
                "角度公差轉換距離" if lang != 'en' else "Ang Tol Dist": item.get("dist", "") or "", 
            }) 

    df = pd.DataFrame(rows) 
    output = io.StringIO() 
    df.to_csv(output, index=False, encoding="utf-8-sig") 
    csv_content = output.getvalue() 

    return Response(
        csv_content,
        mimetype="text/csv",
        headers={"Content-disposition": "attachment; filename=Tolerance_Path_Export.csv"}, 
    ) 


@app.route("/api/export_contact_lines_csv", methods=["POST"])
def export_contact_lines_csv():
    data = request.get_json() 
    pairs = data.get("pairs", []) 
    lang = data.get("lang", "zh-TW") 

    rows = [] 
    for pair in pairs:
        rows.append({
            "特徵面 1" if lang != 'en' else "Feature 1": pair.get("start"), 
            "特徵面 2" if lang != 'en' else "Feature 2": pair.get("end"), 
            "連結類型" if lang != 'en' else "Connection Type": "硬接觸" if lang != 'en' else "Hard Contact", 
        }) 

    df = pd.DataFrame(rows) 
    output = io.StringIO() 
    df.to_csv(output, index=False, encoding="utf-8-sig") 
    csv_content = output.getvalue() 

    return Response(
        csv_content,
        mimetype="text/csv",
        headers={"Content-disposition": "attachment; filename=Contact_Lines_Export.csv"}, 
    ) 


@app.route("/api/sync_report", methods=["POST"])
def sync_report():
    try:
        data = request.get_json() 
        report_text = data.get("reportText", "") 

        if not report_text:
            return jsonify({"ok": False, "msg": "沒有收到報表內容"}), 400 

        graph_rag.set_latest_report(report_text) 
        print(f"[SUCCESS] 成功接收並更新最新機台媒合報表 (長度: {len(report_text)})") 
        return jsonify({"ok": True, "msg": "報表同步成功"}) 

    except Exception as e:
        return jsonify({"ok": False, "msg": f"同步失敗: {str(e)}"}), 500 


@app.route("/api/export_analysis_excel", methods=["POST"])
def export_analysis_excel_route():
    """接收分析結果並回傳格式化後的專業 Excel 報表"""
    body = request.get_json(force=True)
    path_data = body.get("pathData", [])
    result    = body.get("result",    {})
    lang      = body.get("lang",      "zh-TW")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Tolerance_Report"
    
    # ── 寫入公差路徑 (Col A-D) ────────────────────────────────────────────────
    hdr = ["Tolerance Code", "Value", "Bias", "Dist"] if lang == "en" else ["路徑代碼", "數值", "偏差", "轉換距離"]
    for col_idx, h in enumerate(hdr, 1):
        c = ws.cell(row=1, column=col_idx)
        c.value = h
        c.font = Font(name='Times New Roman', size=11, bold=True)
    
    for r_idx, item in enumerate(path_data, 2):
        code = item.get('name') if item.get('type') == 'feature' else item.get('axis', '')
        ws.cell(row=r_idx, column=1).value = code
        ws.cell(row=r_idx, column=2).value = item.get('val', 0)
        ws.cell(row=r_idx, column=3).value = item.get('bias', 0)
        ws.cell(row=r_idx, column=4).value = item.get('dist', 1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Tolerance_Analysis_Report.xlsx"},
    )


# ── [重構] 全域 Excel 樣式與工具 ──────────────────────────────────────────────
RAD_TO_ARCSEC = 206264.8
PINK_FILL   = PatternFill(start_color='fbcfe8', end_color='fbcfe8', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='fef08a', end_color='fef08a', fill_type='solid')
GREEN_FILL  = PatternFill(start_color='bbf7d0', end_color='bbf7d0', fill_type='solid')
TAN_FILL    = PatternFill(start_color='bd9a7a', end_color='bd9a7a', fill_type='solid')
GREY_FILL   = PatternFill(start_color='94a3b8', end_color='94a3b8', fill_type='solid')
BLUE_FILL   = PatternFill(start_color='bfdbfe', end_color='bfdbfe', fill_type='solid') # Sensitivity
RED_FILL    = PatternFill(start_color='fecaca', end_color='fecaca', fill_type='solid') # Contribution
DBLUE_FILL  = PatternFill(start_color='93c5fd', end_color='93c5fd', fill_type='solid') # Angle Sens
MAUVE_FILL  = PatternFill(start_color='e2e8f0', end_color='e2e8f0', fill_type='solid') # Angle Cont

FONT_TNR      = Font(name='Times New Roman', size=10)
FONT_TNR_BOLD = Font(name='Times New Roman', size=11, bold=True)

def _write_prof_block(ws, start_row, start_col, title, fill, headers, rows):
    cell = ws.cell(row=start_row, column=start_col)
    cell.value = title
    cell.font = FONT_TNR_BOLD
    ws.merge_cells(start_row=start_row, start_column=start_col, 
                   end_row=start_row, end_column=start_col + len(headers) - 1)
    if headers:
        for i, h in enumerate(headers):
            if h:
                c = ws.cell(row=start_row + 1, column=start_col + i)
                c.value = h
                c.font = FONT_TNR
    for r_idx, r_data in enumerate(rows):
        for c_idx, val in enumerate(r_data):
            c = ws.cell(row=start_row + 2 + r_idx, column=start_col + c_idx)
            c.value = val
            c.font = FONT_TNR
    last_row = start_row + 2 + len(rows) - 1
    last_col = start_col + len(headers) - 1
    for r in range(start_row, last_row + 1):
        for col in range(start_col, last_col + 1):
            ws.cell(row=r, column=col).fill = fill

def _write_prof_summary_blocks(ws, result):
    """共同的專業分析塊寫入 (Tideal, RSS, WC, SCA)"""
    # 寫入摘要區塊
    tm = result.get("t_ideal_matrix", [[1,0,0,0],[0,1,0,0],[0,0,1,0],[0,0,0,1]])
    _write_prof_block(ws, 1, 11, "Tideal Matrix", PINK_FILL, ["", "", "", ""], [
        [tm[0][0], tm[0][1], tm[0][2], tm[0][3]],
        [tm[1][0], tm[1][1], tm[1][2], tm[1][3]],
        [tm[2][0], tm[2][1], tm[2][2], tm[2][3]],
        [tm[3][0], tm[3][1], tm[3][2], tm[3][3]]
    ])
    _write_prof_block(ws, 8, 11, "Statistics Model(RSS)", YELLOW_FILL, ["tol_range", "-3sigma", "+3sigma"], [
        ["Xerror", result.get("rss_X",0)*-1, result.get("rss_X",0)],
        ["Yerror", result.get("rss_Y",0)*-1, result.get("rss_Y",0)],
        ["Zerror", result.get("rss_Z",0)*-1, result.get("rss_Z",0)]
    ])
    _write_prof_block(ws, 14, 11, "Worst Case Model", GREEN_FILL, ["tol_range", "min", "max"], [
        ["Xerror", result.get("wc_X",0)*-1, result.get("wc_X",0)],
        ["Yerror", result.get("wc_Y",0)*-1, result.get("wc_Y",0)],
        ["Zerror", result.get("wc_Z",0)*-1, result.get("wc_Z",0)]
    ])
    # SCA 表格
    def _get_table_rows(data_list):
        return [[d.get('name',''), d.get('x',0), d.get('y',0), d.get('z',0)] for d in data_list[:8]]

    _write_prof_block(ws, 1, 16, "Sensitivity Analysis", BLUE_FILL, ["tol_sym", "X(%)", "Y(%)", "Z(%)"], _get_table_rows(result.get("sensitivity", [])))
    _write_prof_block(ws, 1, 22, "Contribution Analysis", RED_FILL, ["tol_sym", "X(%)", "Y(%)", "Z(%)"], _get_table_rows(result.get("contribution", [])))


@app.route("/api/export_allocation_excel", methods=["POST"])
def export_allocation_excel():
    """產出專業級公差調配比對報表 (包含前後對比與優化後分析)"""
    body = request.get_json(force=True)
    prev_path = body.get("prevPathData", [])
    new_path = body.get("newPathData", [])
    report = body.get("report", {})
    new_result = body.get("analysisResult", {}) # 優化後的完整分析結果
    lang = body.get("lang", "zh-TW")

    wb = Workbook()
    
    # ── [Sheet 1] 調配前後對比清單 ──
    ws1 = wb.active
    ws1.title = "Allocation_Comparison"
    
    # RSS 改善摘要 (頂部區塊)
    ws1.cell(row=1, column=1).value = "【RSS 預測改善匯總】" if lang != 'en' else "[RSS Improvement Summary]"
    ws1.cell(row=1, column=1).font = FONT_TNR_BOLD
    
    hdr_rss = ["軸向", "目前 (Before)", "優化後 (After)", "改善 (%)"] if lang != 'en' else ["Axis", "Before", "After", "Improve %"]
    for i, h in enumerate(hdr_rss, 1):
        ws1.cell(row=2, column=i).value = h
        ws1.cell(row=2, column=i).fill = YELLOW_FILL
    
    axes = ['X', 'Y', 'Z', 'aX', 'aY', 'aZ']
    curr_r = 3
    for ax in axes:
        item = report.get(ax)
        if not item: continue
        ws1.cell(row=curr_r, column=1).value = ax
        ws1.cell(row=curr_r, column=2).value = item.get('rss_before', 0)
        ws1.cell(row=curr_r, column=3).value = item.get('rss_after', 0)
        ws1.cell(row=curr_r, column=4).value = f"{item.get('rss_improve_pct', 0)}%"
        curr_r += 1

    # 公差詳細變動表 (下方區塊)
    curr_r += 2
    ws1.cell(row=curr_r, column=1).value = "【個別公差明細表】" if lang != 'en' else "[Tolerance Detail Table]"
    ws1.cell(row=curr_r, column=1).font = FONT_TNR_BOLD
    curr_r += 1
    
    hdr_tol = ["項目", "原公差", "新公差", "變動 (%)", "診斷象限"] if lang != 'en' else ["Item", "Old Tol", "New Tol", "Delta %", "Quadrant"]
    for i, h in enumerate(hdr_tol, 1):
        ws1.cell(row=curr_r, column=i).value = h
        ws1.cell(row=curr_r, column=i).fill = BLUE_FILL
    
    prev_map = {p.get('name'): p.get('val', 0) for p in prev_path if p.get('name')}
    q_names = {1:'Q1(關鍵)', 2:'Q2(維護)', 3:'Q3(次要)', 4:'Q4(放寬)'} if lang != 'en' else {1:'Q1', 2:'Q2', 3:'Q3', 4:'Q4'}
    
    for item in new_path:
        if item.get('type') != 'feature': continue
        curr_r += 1
        name = item.get('name')
        new_val = item.get('val', 0)
        old_val = prev_map.get(name, new_val)
        delta = ((new_val - old_val)/old_val*100) if old_val > 0 else 0
        q = item.get('quadrant', 4)
        
        ws1.cell(row=curr_r, column=1).value = name
        ws1.cell(row=curr_r, column=2).value = old_val
        ws1.cell(row=curr_r, column=3).value = new_val
        ws1.cell(row=curr_r, column=4).value = f"{round(delta, 1)}%"
        ws1.cell(row=curr_r, column=5).value = q_names.get(q, 'Q4')

    # ── [Sheet 2] 優化後專業分析 (與分析報表規格一致) ──
    ws2 = wb.create_sheet("Optimized_Analysis")
    # 寫入路徑
    for col_idx, h in enumerate(hdr_rss, 1): # 複用 hdr
        ws2.cell(row=1, column=col_idx).value = h
    for r_idx, item in enumerate(new_path, 2):
        code = item.get('name') if item.get('type') == 'feature' else item.get('axis', '')
        ws2.cell(row=r_idx, column=1).value = code
        ws2.cell(row=r_idx, column=2).value = item.get('val', 0)
        ws2.cell(row=r_idx, column=3).value = item.get('bias', 0)
        ws2.cell(row=r_idx, column=4).value = item.get('dist', 1)
    
    # 寫入專業分析塊
    _write_prof_summary_blocks(ws2, new_result)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Tolerance_Allocation_Report_V{round}.xlsx"}
    )


@app.route("/api/run_allocation", methods=["POST"])
def run_allocation():
    """
    公差調配 API。
    支援兩種模式：
      mode='auto'   → 依目標 RSS 與權重自動計算建議公差值
      mode='compare'→ 手動調配比對，回傳改善百分比報告
    """
    body           = request.get_json(force=True)
    mode           = body.get("mode", "auto")           # 'auto' | 'compare'
    path_data      = body.get("pathData", [])
    analysis_result= body.get("analysisResult", {})
    baseline       = body.get("baseline", {})

    try:
        from analysis_service import compute_allocation, compare_allocation

        if mode == "auto":
            target = float(body.get("target", 0.05))
            strategy = body.get("weight", "medium")    # 前端傳遞 'weight'，後端解讀為 strategy
            axis   = body.get("axis", "Z")
            res_data = compute_allocation(path_data, analysis_result, target, strategy, axis)
            
            # [新增] 為左側面板生成更新後的 DSL
            # 提取所有特徵的目標值與象限
            overrides = {}
            quadrants = {}
            report = res_data.get("report", {})
            q_map = report.get("quadrants", {})
            
            for item in res_data["newPathData"]:
                if item.get("type") == "feature":
                    name = item.get("name")
                    overrides[name] = item.get("val", 0)
                    if name in q_map:
                        quadrants[name] = q_map[name]

            updated_dsl = build_full_dsl(mode='network', tolerance_overrides=overrides, quadrants=quadrants)
            
            # [核心] 在回傳前，對新數據執行一次完整分析，以便產出專業報表
            import analysis_service
            final_res = analysis_service.analyze_tolerance_path(res_data["newPathData"])

            # [新增] 生成完整 6 軸向對比報告 (包含 RSS 與 WorstCase)
            full_report = {}
            for ax in ['X', 'Y', 'Z', 'aX', 'aY', 'aZ']:
                rb = analysis_result.get(f'rss_{ax}', 0) or 0
                ra = final_res.get(f'rss_{ax}', 0) or 0
                wb = analysis_result.get(f'wc_{ax}', 0) or 0
                wa = final_res.get(f'wc_{ax}', 0) or 0
                
                def _ipct(b, a):
                    if b == 0: return 0.0
                    return round((b - a) / b * 100, 2)

                full_report[ax] = {
                    'rss_before': round(rb, 6),
                    'rss_after':  round(ra, 6),
                    'rss_improve_pct': _ipct(rb, ra),
                    'wc_before':  round(wb, 6),
                    'wc_after':   round(wa, 6),
                    'wc_improve_pct':  _ipct(wb, wa)
                }
            full_report['quadrants'] = q_map

            return jsonify({
                "ok": True, 
                "mode": "auto", 
                "newPathData": res_data["newPathData"],
                "report": full_report,
                "dsl": updated_dsl,
                "analysisResult": final_res
            })

        elif mode == "compare":
            # [核心修正] 手動調配 (單機版比對) 模式：按下開始後先執行最新分析，再進行比對
            import analysis_service
            current_analysis = analysis_service.analyze_tolerance_path(path_data)
            
            res_data = compare_allocation(baseline, current_analysis)
            
            # [新增] 在比對模式中也加入象限分析，讓手動調整也能看到診斷結果
            names = [item['name'] for item in path_data if item.get('type') == 'feature']
            q_map, _ = analysis_service.get_quadrant_analysis(current_analysis, names, body.get('axis', 'Z'))
            res_data['quadrants'] = q_map
            
            # 將象限資訊注入 path_data 以供前端表格顯示
            updated_path = copy.deepcopy(path_data)
            for item in updated_path:
                if item.get('type') == 'feature' and item.get('name') in q_map:
                    item['quadrant'] = q_map[item['name']]

            return jsonify({
                "ok": True, 
                "mode": "compare", 
                "report": res_data,
                "newPathData": updated_path, # [新增] 同步帶有象限資訊的路徑
                "analysisResult": current_analysis # 回傳最新分析結果給前端同步
            })

        else:
            return jsonify({"ok": False, "error": f"未知模式: {mode}"}), 400

    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/api/save_allocation", methods=["POST"])
def save_allocation_to_excel():
    """將調配後的數據存入 Excel 的新分頁中 (Allocation History)"""
    body = request.get_json(force=True)
    path_data = body.get("pathData", [])
    result    = body.get("result",    {})
    lang      = body.get("lang",      "zh-TW")
    
    excel_path = "test.xlsx"
    if os.path.exists(excel_path):
        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames: del wb["Sheet"]

    # 計算下一個分頁名稱
    alloc_sheets = [s for s in wb.sheetnames if s.startswith("Allocation_V")]
    next_ver = len(alloc_sheets) + 1
    sheet_name = f"Allocation_V{next_ver}"
    
    ws = wb.create_sheet(sheet_name)
    
    # 寫入路徑數據
    hdr = ["Path Code", "Value", "Bias", "Dist"] if lang == "en" else ["路徑代碼", "數值", "偏差", "轉換距離"]
    for col, h in enumerate(hdr, 1):
        ws.cell(row=1, column=col).value = h
    
    for i, item in enumerate(path_data, 2):
        code = item.get('name') if item.get('type') == 'feature' else item.get('axis', '')
        ws.cell(row=i, column=1).value = code
        ws.cell(row=i, column=2).value = item.get('val', 0)
        ws.cell(row=i, column=3).value = item.get('bias', 0)
        ws.cell(row=i, column=4).value = item.get('dist', 1)

    # 寫入簡單摘要
    ws.cell(row=1, column=6).value = "Analysis Results"
    ws.cell(row=2, column=6).value = "RSS (±3σ)"
    ws.cell(row=2, column=7).value = result.get("rss_X_std", 0) * 3
    
    try:
        wb.save(excel_path)
        return jsonify({"ok": True, "msg": f"Saved to {sheet_name}"})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500


# ═══════════════════════════════════════════════════════════
# STEP PMI 路由註冊 (Phase 1 整合)
# ═══════════════════════════════════════════════════════════
app.add_url_rule('/api/step/upload',      'step_upload',     route_upload_step,     methods=['POST'])
app.add_url_rule('/api/step/parse_pmi',   'step_parse_pmi',  route_parse_pmi,       methods=['POST'])
app.add_url_rule('/api/step/geometry',    'step_geometry',   route_get_geometry,    methods=['GET'])
app.add_url_rule('/api/step/pmi_list',    'step_pmi_list',   route_get_pmi_list,    methods=['GET'])
app.add_url_rule('/api/step/highlight',   'step_highlight',  route_highlight_pmi,   methods=['POST'])
app.add_url_rule('/api/step/asm_contact', 'step_asm_contact',route_run_asm_worker,  methods=['POST'])
app.add_url_rule('/api/step/6dof',        'step_6dof',       route_get_6dof,        methods=['POST'])
app.add_url_rule('/api/step/export_csv',  'step_export_csv', route_export_step_csv, methods=['POST'])

# 進度追蹤路由 (新增)
app.add_url_rule('/api/step/progress',    'step_progress',   route_progress_sse,    methods=['GET'])
app.add_url_rule('/api/step/progress_status', 'step_progress_status', route_progress_status, methods=['GET'])


if __name__ == "__main__":
    print("啟動 AI 聊天助手伺服器...")
    print("請訪問: http://127.0.0.1:7011")
    app.run(host="0.0.0.0", port=7011, debug=True)

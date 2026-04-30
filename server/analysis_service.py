# -*- coding: utf-8 -*-
"""
analysis_service.py — Web 服務層，銜接前端 JSON 路徑與分析引擎。
負責：
  1. 將前端 editorPathData JSON 轉換為 ToleranceData 物件。
  2. 以 Thread + Queue 執行 Jacobian 分析，透過 SSE generator 推播進度。
"""
import json
import queue
import threading
import math
import copy
import numpy as np

from analysis_engine.jacobian   import compute_jacobian
from analysis_engine.statistics import compute_rss, compute_monte_carlo, compute_sensitivity_contribution

# 軸向與引擎結果 Key 的映射關係
axis_key_map = {
    'X': 'sens_X', 'Y': 'sens_Y', 'Z': 'sens_Z',
    'aX': 'sens_aX', 'aY': 'sens_aY', 'aZ': 'sens_aZ'
}


# ─── 角度類公差判斷（需除以 dist 轉換量綱，對應 ExcelData.py 行為）──────────────
_ANGULAR_PATTERNS = ('ang', 'per', 'par', 'cra')

def _is_angular(name: str) -> bool:
    nl = name.lower()
    return any(p in nl for p in _ANGULAR_PATTERNS)


# ─── 資料模型 ──────────────────────────────────────────────────────────────────

class ToleranceData:
    """
    從前端 pathData JSON 建立供 Jacobian 引擎使用的資料容器。

    前端格式：
      [{"type": "spatial", "axis": "traX", "val": 10.0, "bias": 0, "dist": 1, "nominal_size": 25.0, "it_grade": "IT7"},
       {"type": "feature", "name": "1-disX-1", "val": 0.02, "bias": 0, "dist": 1}]
    """

    def __init__(self, path_items: list):
        self.tolnum    = len(path_items)
        self.all_sym   = []
        self.all_value = []
        self.tol_names = []
        self.tol_values = []
        self.tol_dis    = []
        self.tol_biases = []   # C 欄：偏差值（距離公差不對稱帶中心）
        # [新增] 存儲原始元數據，用於更新與回傳
        self.nominal_sizes = []
        self.it_grades = []

        # 只有純空間平移/旋轉才算 spatial
        PURE_SPATIAL_AXES = {'traX', 'traY', 'traZ', 'rotX', 'rotY', 'rotZ'}

        for j, item in enumerate(path_items):
            raw_val = float(item.get('val', 0.0))
            self.all_value.append(raw_val)
            self.nominal_sizes.append(item.get('nominal_size'))
            self.it_grades.append(item.get('it_grade'))

            axis = item.get('axis', '')
            is_pure_spatial = (item.get('type') == 'spatial') and (axis in PURE_SPATIAL_AXES)

            if is_pure_spatial:
                self.all_sym.append(axis)
            else:
                # feature 項目
                name = item.get('name') or axis or f'feat_{j}'
                self.all_sym.append(name)
                self.tol_names.append(name)

                # D 欄：角度公差轉換距離，預設 1
                dist = float(item.get('dist') or 1.0)
                if dist == 0:
                    dist = 1.0
                self.tol_dis.append(dist)

                # C 欄：偏差值（距離公差不對稱帶的中心偏移）
                # 幾何公差（per/fla/par/cir 等）理論上 bias=0；
                # 距離公差（dis）才需要使用 bias。
                raw_bias = float(item.get('bias', 0.0))
                tol_type = item.get('tol_type', '')
                # 非距離公差的 bias 強制歸零（遵循說明書規定）
                if tol_type and tol_type != 'dis':
                    bias_val = 0.0
                else:
                    bias_val = raw_bias
                self.tol_biases.append(bias_val)

                # 角度類公差值預除 dist
                if _is_angular(name):
                    eff_val = raw_val / dist
                else:
                    eff_val = raw_val
                self.tol_values.append(eff_val)

        # 分析結果（由引擎填入）
        self.sens_X  = []
        self.sens_Y  = []
        self.sens_Z  = []
        self.sens_aX = []
        self.sens_aY = []
        self.sens_aZ = []


from openpyxl import load_workbook
import io
import csv as _csv

# 公差軟體說明書的標準 XLSX 欄位順序：
#   A=代碼, B=值, C=偏差, D=距離(角度), E=公稱尺寸, F=IT等級
#
# 也支援使用者直接拖 skeleton_chain.csv，header 可能變體很多：
#   - 路徑代碼, 公稱尺寸, IT等級, 數值, 偏差值, 角度公差
#   - A 路徑代碼, E 公稱尺寸, F IT 等級, B 數值, C 偏差值, D 角度公差   ← 帶 Excel 字母前綴
#   - 只要 header 含關鍵字皆能識別

PURE_SPATIAL_AXES = {'traX', 'traY', 'traZ', 'rotX', 'rotY', 'rotZ'}


def _norm_header(s):
    """正規化 header 字串：去前後空白、移除空格、去除 'A '/'B '/.../ Excel 字母前綴。"""
    if not s:
        return ''
    s = s.strip().replace(' ', '').replace('　', '')   # 移除半形+全形空格
    # 去除常見前綴 A/B/C/D/E/F（如 'A路徑代碼' 變 '路徑代碼'）
    if len(s) >= 2 and s[0] in 'ABCDEF' and not s[1].isalnum():
        s = s[1:].lstrip()
    elif len(s) >= 2 and s[0] in 'ABCDEF':
        # 如 'B數值' / 'A路徑代碼' (大寫前綴緊接著中文/IT)
        rest = s[1:]
        if rest and (not rest[0].isascii() or rest.startswith('IT')):
            s = rest
    return s


def _build_skeleton_col_map(headers):
    """掃 header → 回傳 {name/nominal/it/val/bias/dist: column_index}。

    用包含關鍵字的方式比對，可容忍 'A 路徑代碼'、'F IT 等級'、'IT等級' 等變體。
    """
    # 比對表：規範 key → 可能含的關鍵詞列表
    matchers = {
        'name':    ['路徑代碼', '代碼', 'sym'],
        'nominal': ['公稱尺寸', '公稱', 'nominal'],
        'it':      ['IT等級', 'IT級', 'IT'],
        'val':     ['數值', '公差值', 'val'],
        'bias':    ['偏差值', '偏差', 'bias'],
        'dist':    ['角度公差', '距離', '角度', 'dist'],
    }
    col_map = {}
    used = set()
    for key, candidates in matchers.items():
        for i, h in enumerate(headers):
            if i in used:
                continue
            nh = _norm_header(h).upper()
            if any(c.upper() in nh for c in candidates):
                col_map[key] = i
                used.add(i)
                break
    return col_map


def _is_skeleton_csv_header(header):
    """偵測是否為 skeleton_chain.csv 標頭格式（含字母前綴變體）。"""
    if not header:
        return False
    col_map = _build_skeleton_col_map(header)
    # 至少要找到 name + (val 或 數值) + (公稱 或 IT) — 三項以上就算
    required = {'name', 'val'}
    optional_at_least_one = {'nominal', 'it', 'bias', 'dist'}
    return required.issubset(col_map.keys()) and bool(set(col_map.keys()) & optional_at_least_one)


def _parse_skeleton_csv(text):
    """解析 skeleton_chain.csv → editorPathData 結構。"""
    items: list[dict] = []
    skipped: list[str] = []
    reader = _csv.reader(io.StringIO(text))
    headers = next(reader, None)
    if not headers:
        raise ValueError("CSV 無 header")
    col_map = _build_skeleton_col_map(headers)
    if 'name' not in col_map or 'val' not in col_map:
        raise ValueError(f"CSV 缺少必要欄位（路徑代碼/數值），header: {headers}")

    def _cell(row, key, default=None):
        idx = col_map.get(key)
        if idx is None or idx >= len(row):
            return default
        v = (row[idx] or '').strip()
        return v if v else default

    def _num(row, key, default=None):
        v = _cell(row, key, default)
        if v in (None, ''):
            return default
        try:
            return float(v)
        except (ValueError, TypeError):
            return v

    for row in reader:
        if not row or not (row[col_map['name']] or '').strip():
            continue
        sym = row[col_map['name']].strip()

        # IT 等級檢查：只接受數字或 ITn；其他 (H/K/L 等) → 留空
        it_raw = _cell(row, 'it', '') or ''
        if it_raw:
            if it_raw.isdigit():
                it_grade = f"IT{it_raw}"
            elif it_raw.upper().startswith('IT') and it_raw[2:].isdigit():
                it_grade = it_raw.upper()
            else:
                skipped.append(f"{sym} (IT={it_raw} → 空)")
                it_grade = None
        else:
            it_grade = None

        val      = _num(row, 'val', 0.0)      or 0.0
        bias     = _num(row, 'bias', 0.0)     or 0.0
        dist     = _num(row, 'dist', 1.0)     or 1.0
        if dist == 0:
            dist = 1.0
        nominal  = _num(row, 'nominal')

        is_spatial = sym in PURE_SPATIAL_AXES
        item = {
            'type':         'spatial' if is_spatial else 'feature',
            'val':          val, 'bias': bias, 'dist': dist,
            'nominal_size': nominal, 'it_grade': it_grade,
        }
        if is_spatial:
            item['axis'] = sym
        else:
            item['name'] = sym
        items.append(item)

    if skipped:
        print(f"[CSV import] {len(skipped)} 列 IT 欄非數字，已留空（公差值仍保留）：{', '.join(skipped[:5])}"
              + (f" ...等共 {len(skipped)} 列" if len(skipped) > 5 else ""))

    if not items:
        raise ValueError("CSV 中找不到有效的公差路徑資料列")
    return items


def parse_excel_to_path(file_stream, filename=None):
    """
    從上傳的檔案流解析公差路徑。

    支援兩種格式（自動偵測）：
      1. .xlsx  — 標準格式：A=代碼, B=值, C=偏差, D=距離, E=公稱, F=IT
      2. .csv   — skeleton_chain.csv：路徑代碼, 公稱尺寸, IT等級, 數值, 偏差值, 角度公差
    """
    # 先嘗試 CSV：副檔名為 .csv，或 XLSX 解析失敗再 fallback
    is_csv = bool(filename and filename.lower().endswith('.csv'))
    if not is_csv:
        # 內容嗅探：xlsx 是 zip（PK\x03\x04），純文字 CSV 開頭通常是字元
        head = file_stream[:4] if isinstance(file_stream, (bytes, bytearray)) else b''
        is_csv = not head.startswith(b'PK')

    if is_csv:
        try:
            text = file_stream.decode('utf-8-sig') if isinstance(file_stream, (bytes, bytearray)) else file_stream
        except UnicodeDecodeError:
            text = file_stream.decode('big5', errors='replace')
        # peek header 確認是 skeleton 格式
        first_line = text.split('\n', 1)[0]
        header = [h.strip() for h in first_line.split(',')]
        print(f"[CSV import] filename={filename!r}  header={header}")
        if _is_skeleton_csv_header(header):
            print(f"[CSV import] → 走 skeleton 6-欄 parser")
            return _parse_skeleton_csv(text)
        print(f"[CSV import] → 走舊 4-欄 fallback parser（headers 不匹配 skeleton 格式）")
        # 一般 CSV：第一欄 sym, 第二欄 val, 第三欄 bias, 第四欄 dist
        # （與 XLSX 欄序相同的單純 CSV）
        items: list[dict] = []
        for row in _csv.reader(io.StringIO(text)):
            if not row or not row[0].strip():
                continue
            sym = row[0].strip()
            try:
                val  = float(row[1]) if len(row) > 1 and row[1] else 0.0
                bias = float(row[2]) if len(row) > 2 and row[2] else 0.0
                dist = float(row[3]) if len(row) > 3 and row[3] else 1.0
            except ValueError:
                continue
            if dist == 0:
                dist = 1.0
            is_spatial = sym in PURE_SPATIAL_AXES
            item = {
                'type':         'spatial' if is_spatial else 'feature',
                'val': val, 'bias': bias, 'dist': dist,
                'nominal_size': None, 'it_grade': None,
            }
            if is_spatial:
                item['axis'] = sym
            else:
                item['name'] = sym
            items.append(item)
        if not items:
            raise ValueError("CSV 中找不到有效的公差路徑資料列")
        return items

    try:
        wb = load_workbook(io.BytesIO(file_stream), data_only=True)
    except Exception as e:
        raise ValueError(f"無法讀取 Excel 檔案，請確認格式是否正確 (.xlsx)。內容解析錯誤: {str(e)}")

    # 嘗試尋找含有公差設定的分頁
    # 邏輯：優先檢查當前分頁，若失敗則檢查所有分頁，尋找 A2 或 A1 有資料的分頁
    target_sheet = None
    # 常用關鍵字分頁優先
    for name in wb.sheetnames:
        if any(k in name.lower() for k in ('tolerance', 'path', '公差', '路徑')):
            target_sheet = wb[name]
            break
            
    if not target_sheet:
        target_sheet = wb.active

    def try_parse_sheet(sheet):
        items = []
        # 從第二行開始讀取 (跳過標題) 或直接從第一行開始 (視第一格內容而定)
        start_row = 1 if sheet['A1'].value and str(sheet['A1'].value).strip() in PURE_SPATIAL_AXES else 2
        
        for row_idx in range(start_row, min(sheet.max_row, 1000) + 1): # 限制 1000 行
            try:
                sym_val = sheet.cell(row=row_idx, column=1).value
                if sym_val is None: 
                    if row_idx > start_row + 2: break # 連續空行則停止
                    continue
                
                sym_str = str(sym_val).strip()
                if not sym_str: continue

                val = float(sheet.cell(row=row_idx, column=2).value or 0.0)
                bias = float(sheet.cell(row=row_idx, column=3).value or 0.0)
                dist = float(sheet.cell(row=row_idx, column=4).value or 1.0)
                if dist == 0: dist = 1.0

                # [新增] 讀取公稱尺寸與 IT 等級
                nominal  = sheet.cell(row=row_idx, column=5).value
                it_grade = sheet.cell(row=row_idx, column=6).value
                # 正規化 it_grade 為 string-or-None（前端會做 .trim()，不可為 int/float）
                if it_grade is None:
                    pass
                elif isinstance(it_grade, (int, float)):
                    it_grade = f"IT{int(it_grade)}"
                else:
                    it_grade = str(it_grade).strip() or None
                    if it_grade and it_grade.isdigit():
                        it_grade = f"IT{it_grade}"

                is_spatial = (sym_str in PURE_SPATIAL_AXES)
                item = {
                    "type": "spatial" if is_spatial else "feature",
                    "val": val, "bias": bias, "dist": dist,
                    "nominal_size": nominal, "it_grade": it_grade
                }
                if is_spatial: item["axis"] = sym_str
                else: item["name"] = sym_str
                items.append(item)
            except (ValueError, TypeError):
                continue
        return items

    path_items = try_parse_sheet(target_sheet)
    
    # 如果選定的分頁沒掃到東西，則掃描所有分頁
    if not path_items:
        for name in wb.sheetnames:
            if wb[name] == target_sheet: continue
            path_items = try_parse_sheet(wb[name])
            if path_items: break

    if not path_items:
        raise ValueError("在 Excel 中找不到有效的公差路徑配置。請確認檔案內容是否符合格式（A欄：代碼，B欄：公差值，C欄：偏差，D欄：角度轉換距離）。")
        
    return path_items


# ─── SSE 串流分析 ──────────────────────────────────────────────────────────────

# 蒙地卡羅原始樣本回傳上限：避免 SSE 大 JSON 寫入時 broken pipe / 記憶體峰值
# 樣本仍照 mc_samples 跑統計（標準差、Worst Case 等準確），只是 mc_raw 只截前 MC_RAW_CAP 筆。
MC_RAW_CAP = 2000


def analyze_stream(path_data: list, run_mc: bool = True, mc_samples: int = 10000, mc_sigma: float = 3.0, mc_dist: int = 0):
    """
    Generator，以 SSE 格式 yield 進度與結果。

    [v3] 同步執行，避免 Windows + conda numpy/MKL + threading 組合的 segfault。
    每個分析階段完成後 yield 一次進度。
    """

    def _safe_yield(payload_dict):
        """Yield 一條 SSE 訊息，若客戶端斷線則靜默返回。"""
        try:
            return f"data: {json.dumps(payload_dict, default=str)}\n\n"
        except Exception:
            return f"data: {json.dumps({'error': 'JSON 序列化失敗'})}\n\n"

    try:
        # 階段 0：建立公差資料物件
        try:
            tol_data = ToleranceData(path_data)
        except Exception as e:
            import traceback
            yield _safe_yield({'error': f'ToleranceData 建立失敗: {e}\n{traceback.format_exc()}'})
            return

        if len(tol_data.tol_names) == 0:
            yield _safe_yield({'error': '路徑中沒有任何公差特徵（feature），無法進行分析。'})
            return

        yield _safe_yield({'progress': 5})

        # 階段 1：Jacobian（不用 progress_cb，避免 thread 互動）
        try:
            t_ideal = compute_jacobian(tol_data)
        except Exception as e:
            import traceback
            yield _safe_yield({'error': f'compute_jacobian 失敗: {e}\n{traceback.format_exc()}'})
            return
        yield _safe_yield({'progress': 60})

        # 階段 2：RSS / Worst Case
        try:
            rss = compute_rss(tol_data)
        except Exception as e:
            import traceback
            yield _safe_yield({'error': f'compute_rss 失敗: {e}\n{traceback.format_exc()}'})
            return
        yield _safe_yield({'progress': 75})

        # 階段 3：蒙地卡羅
        mc = {}
        if run_mc:
            try:
                mc = compute_monte_carlo(tol_data, n_samples=mc_samples, sigma=mc_sigma, dist_type=mc_dist)
                # 截斷 mc_raw 避免 SSE 大 JSON 失敗（統計值仍照完整樣本算）
                if 'mc_raw' in mc and len(mc['mc_raw']) > MC_RAW_CAP:
                    mc['mc_raw'] = mc['mc_raw'][:MC_RAW_CAP]
                    mc['mc_raw_truncated_from'] = mc_samples
            except Exception as e:
                import traceback
                yield _safe_yield({'error': f'compute_monte_carlo 失敗: {e}\n{traceback.format_exc()}'})
                return
        yield _safe_yield({'progress': 90})

        # 階段 4：敏感度 / 貢獻度
        try:
            sc = compute_sensitivity_contribution(tol_data)
        except Exception as e:
            import traceback
            yield _safe_yield({'error': f'compute_sensitivity_contribution 失敗: {e}\n{traceback.format_exc()}'})
            return
        yield _safe_yield({'progress': 99})

        # 階段 5：組合結果
        result = {
            'tol_names':  tol_data.tol_names,
            'tol_values': tol_data.tol_values,
            't_ideal_matrix': t_ideal,
            'sens_X':  tol_data.sens_X,
            'sens_Y':  tol_data.sens_Y,
            'sens_Z':  tol_data.sens_Z,
            'sens_aX': tol_data.sens_aX,
            'sens_aY': tol_data.sens_aY,
            'sens_aZ': tol_data.sens_aZ,
            **rss,
            **mc,
            **sc,
        }
        yield _safe_yield({'result': result})

    except GeneratorExit:
        return
    except Exception as e:
        import traceback
        yield _safe_yield({'error': f'analyze_stream 例外: {e}\n{traceback.format_exc()}'})


def analyze_tolerance_path(path_data: list, mc_samples: int = 10000, mc_sigma: float = 3.0, mc_dist: int = 0):
    """
    同步執行完整公差分析 (Jacobian + RSS + MC + SCA)。
    供 ai_app.py 在調配計算後進行優化結果核實。
    """
    tol_data = ToleranceData(path_data)
    if len(tol_data.tol_names) == 0:
        raise ValueError("路徑中沒有任何公差特徵，無法進行分析。")

    t_ideal = compute_jacobian(tol_data)
    rss = compute_rss(tol_data)
    mc = compute_monte_carlo(tol_data, n_samples=mc_samples, sigma=mc_sigma, dist_type=mc_dist)
    sc = compute_sensitivity_contribution(tol_data)

    return {
        "tol_names":  tol_data.tol_names,
        "tol_values": tol_data.tol_values,
        "t_ideal_matrix": t_ideal,
        "sens_X":  tol_data.sens_X,
        "sens_Y":  tol_data.sens_Y,
        "sens_Z":  tol_data.sens_Z,
        "sens_aX": tol_data.sens_aX,
        "sens_aY": tol_data.sens_aY,
        "sens_aZ": tol_data.sens_aZ,
        **rss,
        **mc,
        **sc
    }



# ─── 公差調配 ──────────────────────────────────────────────────────────────────


def get_quadrant_analysis(analysis_result: dict, tol_names: list, axis: str = 'Z'):
    """計算指定軸向的四象限分析結果 (Q1-Q4)"""
    n = len(tol_names)
    sk = axis_key_map.get(axis, 'sens_Z')
    ck = axis_key_map.get(axis, 'sens_Z').replace('sens', 'ctb')
    
    sens = np.array(analysis_result.get(sk, [0]*n))
    ctb  = np.array(analysis_result.get(ck, [0]*n))
    
    # 標準差與貢獻百分比中位數
    s_mid = np.median(np.abs(sens))
    total_ctb = np.sum(np.abs(ctb))
    c_ratio = (np.abs(ctb) / total_ctb * 100) if total_ctb > 0 else np.zeros(n)
    c_mid = np.median(c_ratio)
    
    quadrants = {}
    for i in range(n):
        si, ci_p = abs(sens[i]), c_ratio[i]
        if si > s_mid and ci_p > c_mid: q = 1
        elif si > s_mid: q = 2
        elif ci_p > c_mid: q = 3
        else: q = 4
        quadrants[tol_names[i]] = q
    return quadrants, c_ratio

def compute_allocation(path_data: list, analysis_result: dict, target: float = 0.05, strategy: str = 'medium', axis: str = 'Z') -> dict:
    """
    自動調配核心邏輯 (原本的 Manual Allocation)
    """
    # 提取特徵名稱與原始值
    names = [item['name'] for item in path_data if item.get('type') == 'feature']
    old_values = np.array([item['val'] for item in path_data if item.get('type') == 'feature'])
    n = len(names)
    
    if n == 0:
        return {"newPathData": path_data, "report": {}}

    sk = axis_key_map.get(axis, 'sens_Z')
    sens = np.array(analysis_result.get(sk, [0]*n))
    
    # 獲取象限與貢獻比 (使用 Helper)
    q_map, cont_ratio_pct = get_quadrant_analysis(analysis_result, names, axis)
    cont_ratio = cont_ratio_pct / 100.0
    
    # 2. 策略權重分配 (w_i)
    w_i = []
    quadrants = []
    for i, name in enumerate(names):
        q = q_map[name]
        quadrants.append(q)
        if strategy == 'precision':
            # 精度優先：針對貢獻度高的項進行更激進的權重係數
            wi = 1.0 - (cont_ratio[i] * 0.8)
        else:
            # 成本優先：針對低敏感 (Q4) 的項放寬
            if q == 4: wi = 1.25
            else: wi = 1.0 - (cont_ratio[i] * 0.5)
        w_i.append(wi)

    # 3. 求解全域因子 β
    t_base = old_values * np.array(w_i)
    weighted_sum_sq = np.sum((sens * t_base)**2)
    
    if weighted_sum_sq < 1e-18:
        return {"newPathData": path_data, "report": {}}
        
    beta = target / math.sqrt(weighted_sum_sq)
    new_values = t_base * beta
    
    # 4. 更新數據與產出報告
    name_to_val = {names[i]: (round(new_values[i], 6), quadrants[i]) for i in range(n)}
    updated = copy.deepcopy(path_data)
    for item in updated:
        if item.get('type') == 'feature':
            name = item.get('name', '')
            if name in name_to_val:
                val, q = name_to_val[name]
                item['val'] = val
                item['quadrant'] = q # 存入象限資訊供前端顯示

    # 預測報告
    report = {}
    axes_list = ['X', 'Y', 'Z', 'aX', 'aY', 'aZ']
    for ax in axes_list:
        sk_ax = axis_key_map.get(ax)
        sl = analysis_result.get(sk_ax, [])
        if not sl or len(sl) != n: continue
        
        pred_rss = math.sqrt(np.sum((np.array(sl) * new_values)**2))
        curr_rss = analysis_result.get(f'rss_{ax}', 0) or 0
        
        def _pct(before, after):
            if before < 1e-12: return 0.0
            return round((before - after) / before * 100, 2)

        report[ax] = {
            'rss_before': round(curr_rss, 6),
            'rss_after':  round(pred_rss, 6),
            'rss_improve_pct': _pct(curr_rss, pred_rss)
        }
    
    report['quadrants'] = q_map
    return {"newPathData": updated, "report": report}


def compare_allocation(baseline: dict, current: dict) -> dict:
    """
    手動調配比對：計算各軸 RSS 與 Worst Case 的改善百分比。
    對應單機版 IntervalAnalysis.py 的 analysis==2 邏輯。

    Returns:
      dict of {axis: {rss_before, rss_after, rss_improve_pct,
                      wc_before, wc_after, wc_improve_pct}}
    """
    axes = ['X', 'Y', 'Z', 'aX', 'aY', 'aZ']
    result = {}
    for ax in axes:
        rss_b = baseline.get(f'rss_{ax}', 0) or 0
        rss_a = current.get(f'rss_{ax}', 0)  or 0
        wc_b  = baseline.get(f'wc_{ax}', 0)  or 0
        wc_a  = current.get(f'wc_{ax}', 0)   or 0

        def _pct(before, after):
            if before == 0:
                return 0.0
            return round((before - after) / before * 100, 2)

        result[ax] = {
            'rss_before':      round(rss_b, 6),
            'rss_after':       round(rss_a, 6),
            'rss_improve_pct': _pct(rss_b, rss_a),
            'wc_before':       round(wc_b, 6),
            'wc_after':        round(wc_a, 6),
            'wc_improve_pct':  _pct(wc_b, wc_a),
        }
    return result

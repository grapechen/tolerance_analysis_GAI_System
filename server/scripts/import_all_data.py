"""完整匯入 ISO 286-1 和 ISO 286-2 資料到 MySQL"""
import pandas as pd
import numpy as np
import re
import openpyxl
from pathlib import Path
from sqlalchemy import text
from sqlalchemy.dialects.mysql import insert as mysql_insert
from tables import Session, ISOTolerance, ShaftTolerance, HoleTolerance

def to_number(x):
    """轉換為數字"""
    if pd.isna(x): 
        return None
    if isinstance(x, (int, float)):
        return float(x)
    if isinstance(x, str):
        s = x.strip().replace('μ','').replace('µ','').replace(',','.')
        s = s.replace('−', '-').replace('–', '-')
        is_negative = '-' in s
        s = re.sub(r'[^\d.]', '', s)
        try:
            num = float(s) if s else None
            if num is not None and is_negative:
                num = -num
            return num
        except:
            return None
    return None

def split_range(text):
    """從文字中提取尺寸範圍"""
    if pd.isna(text): 
        return (None, None)
    s = str(text).strip().replace('－', '0 ').replace('-', ' ')
    nums = re.findall(r'\d+(?:\.\d+)?', s)
    if len(nums) >= 2:
        return (float(nums[0]), float(nums[1]))
    elif len(nums) == 1:
        return (0.0, float(nums[0]))
    return (None, None)

def extract_it(val):
    if pd.isna(val): return None
    s = str(val).strip()
    match = re.match(r'^(\d+)', s)
    if match:
        return int(match.group(1))
    return None

def get_initial_tol(sheet_name):
    match = re.search(r'孔([A-Z]+?)[至的]', sheet_name)
    if match: return match.group(1)
    match = re.search(r'孔([A-Z]+)', sheet_name)
    if match: return match.group(1)
    return None

def import_it_tolerance(excel_path):
    """匯入 IT 基本公差 (ISO 286-1)"""
    print("\n=== 匯入 IT 基本公差 (ISO 286-1) ===")
    try:
        df = pd.read_excel(excel_path, sheet_name='標準公差等級值')
    except:
        print("⚠️ 找不到 '標準公差等級值' 工作表")
        return 0

    df.columns = [str(c).strip() for c in df.columns]
    
    it_cols = [c for c in df.columns if str(c).upper().startswith("IT")]
    non_it_cols = [c for c in df.columns if c not in it_cols]
    
    if len(non_it_cols) < 2:
        print("⚠️ 無法識別尺寸範圍欄位")
        return 0
    
    c_left, c_right = non_it_cols[0], non_it_cols[1]
    range_str = df[c_left].astype(str).fillna('') + ' ' + df[c_right].astype(str).fillna('')
    size_from, size_to = zip(*range_str.map(split_range))
    
    df['size_from_mm'] = pd.Series(size_from, dtype='float64')
    df['size_to_mm'] = pd.Series(size_to, dtype='float64')
    
    for c in it_cols:
        df[c] = df[c].map(to_number)
    
    df = df.dropna(subset=['size_from_mm', 'size_to_mm'])
    
    long_df = df.melt(
        id_vars=['size_from_mm', 'size_to_mm'],
        value_vars=it_cols,
        var_name='it_grade',
        value_name='tolerance_um'
    )
    
    long_df['it_grade'] = long_df['it_grade'].astype(str).str.upper().str.strip()
    long_df['tolerance_um'] = long_df['tolerance_um'].apply(lambda x: to_number(x))
    long_df = long_df.dropna(subset=['tolerance_um'])
    
    inserted = 0
    with Session() as s:
        with s.begin():
            for _, r in long_df.iterrows():
                ins = mysql_insert(ISOTolerance).values(
                    size_from_mm=r['size_from_mm'],
                    size_to_mm=r['size_to_mm'],
                    it_grade=r['it_grade'],
                    tolerance_um=r['tolerance_um'],
                )
                stmt = ins.on_duplicate_key_update(
                    tolerance_um=ins.inserted.tolerance_um
                )
                s.execute(stmt)
                inserted += 1
    
    print(f"✅ IT 公差匯入完成: {inserted} 筆")
    return inserted

def import_hole_tolerance_iso1(excel_path):
    """從 ISO 286-1 讀取孔公差 (如果有的話)"""
    print("  正在讀取 ISO 286-1 孔公差...")
    data = {} # Key: (size_from, size_to, code, it), Value: (upper, lower)
    
    sheets = ['孔A至M的基本偏差值', '孔N至ZC的基本偏差值']
    
    # 檢查工作表是否存在
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        existing_sheets = wb.sheetnames
        wb.close()
    except:
        return {}

    # 預先載入 IT 公差表以加速查詢
    it_map = {} # (size_from, size_to, it_grade) -> tolerance_um
    with Session() as s:
        it_rows = s.query(ISOTolerance).all()
        for row in it_rows:
            it_map[(row.size_from_mm, row.size_to_mm, row.it_grade)] = float(row.tolerance_um)

    for sheet_name in sheets:
        if sheet_name not in existing_sheets:
            continue
            
        try:
            # 跳過前兩行標題
            df = pd.read_excel(excel_path, sheet_name=sheet_name, skiprows=2)
            
            # Handle merged cells (forward fill) for all columns
            df = df.ffill()
            
            # 取得尺寸範圍欄位
            size_cols = df.columns[:2]
            range_str = df[size_cols[0]].astype(str).fillna('') + ' ' + df[size_cols[1]].astype(str).fillna('')
            size_from_list, size_to_list = zip(*range_str.map(split_range))
            
            df['size_from_mm'] = pd.Series(size_from_list, dtype='float64')
            df['size_to_mm'] = pd.Series(size_to_list, dtype='float64')
            
            # 取得公差代號欄位
            tolerance_cols = []
            for col in df.columns[2:]:
                col_str = str(col).strip().upper()
                if 'UNNAMED' in col_str or '(' in col_str and not re.search(r'[A-Z]', col_str): continue
                # 清除註記，例如 N(a)(b) -> N
                clean_code = re.sub(r'\(.*?\)', '', col_str).strip()
                if re.match(r'^[A-Z]+$', clean_code):
                    tolerance_cols.append(col)
            
            for idx, row in df.iterrows():
                if pd.isna(row['size_from_mm']) or pd.isna(row['size_to_mm']):
                    continue
                    
                size_from = round(float(row['size_from_mm']), 1)
                size_to = round(float(row['size_to_mm']), 1)
                
                for col in tolerance_cols:
                    value = to_number(row[col])
                    if value is None: continue
                    
                    # 必須再次清除註記，因為 col 是原始欄位名稱
                    tolerance_code = str(col).strip().upper()
                    tolerance_code = re.sub(r'\(.*?\)', '', tolerance_code).strip()
                    
                    # 針對每個 IT 等級計算上下偏差
                    it_grades_list = ['IT01', 'IT0'] + [f"IT{i}" for i in range(1, 19)]
                    for it_grade in it_grades_list:
                        
                        it_tol = None
                        if (size_from, size_to, it_grade) in it_map:
                            it_tol = it_map[(size_from, size_to, it_grade)]
                        else:
                            mid_size = (size_from + size_to) / 2
                            for (s_f, s_t, ig), tol in it_map.items():
                                if ig == it_grade and s_f <= mid_size and s_t >= mid_size:
                                    it_tol = tol
                                    break
                        
                        if it_tol is None: continue
                        
                        # 判斷是 EI 還是 ES
                        # A-H: 表格值為 EI (Lower). ES = EI + IT
                        # K-ZC: 表格值為 ES (Upper). EI = ES - IT
                        # JS: 對稱. ES = +IT/2, EI = -IT/2
                        # J: 較複雜，暫時略過或需特殊處理
                        
                        is_ei_based = tolerance_code in ['A', 'B', 'C', 'CD', 'D', 'E', 'EF', 'F', 'FG', 'G', 'H']
                        is_es_based = tolerance_code in ['K', 'M', 'N', 'P', 'R', 'S', 'T', 'U', 'V', 'X', 'Y', 'Z', 'ZA', 'ZB', 'ZC']
                        
                        if tolerance_code == 'JS':
                            half_it = it_tol / 2
                            upper_dev = half_it
                            lower_dev = -half_it
                        elif is_ei_based:
                            lower_dev = value
                            upper_dev = lower_dev + it_tol
                        elif is_es_based:
                            upper_dev = value
                            lower_dev = upper_dev - it_tol
                        else:
                            # J 或其他未定義代號
                            continue
                            
                        key = (size_from, size_to, tolerance_code, it_grade)
                        data[key] = (upper_dev, lower_dev)
                        
        except Exception as e:
            print(f"  讀取 ISO 1 {sheet_name} 失敗: {e}")
            
    return data

def import_hole_tolerance_iso2(excel_path):
    """從 ISO 286-2 讀取孔公差"""
    print("  正在讀取 ISO 286-2 孔公差...")
    data = {} 
    if not excel_path.exists(): return data

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
    except: return data

    for sheet_name in sheet_names:
        try:
            df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None, engine='openpyxl')
            if len(df) < 3: continue
            
            row0 = df.iloc[0].tolist()
            row1 = df.iloc[1].tolist()
            row2 = df.iloc[2].tolist()
            col_map = {} 
            
            current_tol = get_initial_tol(sheet_name)
            
            for i in range(2, len(row0)):
                val0 = row0[i]
                if pd.notna(val0):
                    s = str(val0).strip()
                    match = re.match(r'^([A-Z]+)', s, re.IGNORECASE)
                    if match: current_tol = match.group(1).upper()
                
                if not current_tol: continue

                it_num = extract_it(row2[i])
                if it_num is None: it_num = extract_it(row1[i])
                
                if it_num is not None:
                    col_map[i] = (current_tol, f"IT{it_num}")

            r_idx = 3
            while r_idx < len(df) - 1:
                s1 = str(df.iloc[r_idx, 0]).strip()
                s2 = str(df.iloc[r_idx, 1]).strip()
                has_digit = bool(re.search(r'\d', s1)) or bool(re.search(r'\d', s2))
                
                if has_digit:
                    size_val1 = to_number(df.iloc[r_idx, 0])
                    size_val2 = to_number(df.iloc[r_idx, 1])
                    if size_val1 is None and s1 in ['-', '−', '–', '－']: size_val1 = 0.0
                    
                    if size_val1 is not None and size_val2 is not None:
                        size_from = float(size_val1)
                        size_to = float(size_val2)
                        
                        for c_idx, (tol_code, it_grade) in col_map.items():
                            val_upper = to_number(df.iloc[r_idx, c_idx])
                            val_lower = to_number(df.iloc[r_idx+1, c_idx])
                            
                            if val_upper is not None and val_lower is not None:
                                key = (size_from, size_to, tol_code, it_grade)
                                data[key] = (val_upper, val_lower)
                    r_idx += 2
                else:
                    r_idx += 1
        except Exception as e:
            print(f"  讀取 ISO 2 {sheet_name} 失敗: {e}")

    return data

def import_hole_tolerance(iso1_path, iso2_path):
    """匯入並合併孔公差"""
    print("\n=== 匯入孔公差 (合併 ISO 1 & 2) ===")
    data_iso1 = import_hole_tolerance_iso1(iso1_path)
    print(f"  ISO 1 數據量: {len(data_iso1)}")
    data_iso2 = import_hole_tolerance_iso2(iso2_path)
    print(f"  ISO 2 數據量: {len(data_iso2)}")
    
    merged_data = data_iso1.copy()
    merged_data.update(data_iso2)
    print(f"  合併後總數據量: {len(merged_data)}")
    
    total_inserted = 0
    with Session() as s:
        with s.begin():
            for (size_from, size_to, code, it), (upper, lower) in merged_data.items():
                ins = mysql_insert(HoleTolerance).values(
                    size_from_mm=size_from,
                    size_to_mm=size_to,
                    tolerance_code=code,
                    it_grade=it,
                    upper_dev_um=upper,
                    lower_dev_um=lower,
                )
                stmt = ins.on_duplicate_key_update(
                    upper_dev_um=ins.inserted.upper_dev_um,
                    lower_dev_um=ins.inserted.lower_dev_um
                )
                s.execute(stmt)
                total_inserted += 1
    print(f"✅ 孔公差匯入完成: {total_inserted} 筆")
    return total_inserted

def import_shaft_tolerance_iso1(excel_path):
    """匯入軸公差 (ISO 286-1) - 回傳字典"""
    print("\n  正在讀取 ISO 286-1 軸公差...")
    sheets = ['軸 a 至 j 之基本偏差值 ', '軸 k 至 zc 之基本偏差值']
    data = {}
    
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        existing_sheets = wb.sheetnames
        wb.close()
    except:
        print("⚠️ 無法讀取 ISO 286-1 檔案")
        return {}

    # 預先載入 IT Map 以加速查詢
    it_map = {}
    with Session() as s:
        it_rows = s.query(ISOTolerance).all()
        for row in it_rows:
            it_map[(row.size_from_mm, row.size_to_mm, row.it_grade)] = float(row.tolerance_um)

    for sheet_name in sheets:
        if sheet_name not in existing_sheets:
            continue
        try:
            df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
            tolerance_codes = []
            for i, val in enumerate(df.iloc[5, 2:]):
                if pd.isna(val): continue
                code_str = str(val).strip().lower()
                code_clean = re.sub(r'\([^)]*\)', '', code_str).strip()
                if re.match(r'^[a-z]+$', code_clean):
                    tolerance_codes.append((i+2, code_clean))
            
            # Handle merged cells (forward fill) for tolerance columns starting from data rows
            df.iloc[6:, 2:] = df.iloc[6:, 2:].ffill()
            
            data_df = df.iloc[6:].copy()
            data_df.columns = range(len(data_df.columns))
            range_str = data_df[0].astype(str).fillna('') + ' ' + data_df[1].astype(str).fillna('')
            size_from, size_to = zip(*range_str.map(split_range))
            data_df['size_from_mm'] = pd.Series(size_from, dtype='float64', index=data_df.index)
            data_df['size_to_mm'] = pd.Series(size_to, dtype='float64', index=data_df.index)
            
            valid_rows = []
            for idx, row in data_df.iterrows():
                if pd.notna(row['size_from_mm']) and pd.notna(row['size_to_mm']):
                    valid_rows.append(idx)
            data_df = data_df.loc[valid_rows]
            
            for _, row in data_df.iterrows():
                size_from = round(float(row['size_from_mm']), 3)
                size_to = round(float(row['size_to_mm']), 3)
                for col_idx, tolerance_code in tolerance_codes:
                    value = to_number(row[col_idx])
                    if value is None: continue
                    
                    # 判斷是 EI 還是 ES (軸公差)
                    # a-h: 表格值為 es (Upper). ei = es - IT
                    # k-zc: 表格值為 ei (Lower). es = ei + IT
                    # js: 對稱. es = +IT/2, ei = -IT/2
                    # j: 特殊
                    
                    is_es_based = tolerance_code in ['a', 'b', 'c', 'cd', 'd', 'e', 'ef', 'f', 'fg', 'g', 'h']
                    is_ei_based = tolerance_code in ['k', 'm', 'n', 'p', 'r', 's', 't', 'u', 'v', 'x', 'y', 'z', 'za', 'zb', 'zc']
                    
                    it_grades_list = ['IT01', 'IT0'] + [f"IT{i}" for i in range(1, 19)]
                    for it_grade in it_grades_list:
                        
                        it_tol = None
                        if (size_from, size_to, it_grade) in it_map:
                            it_tol = it_map[(size_from, size_to, it_grade)]
                        else:
                            mid_size = (size_from + size_to) / 2
                            for (s_f, s_t, ig), tol in it_map.items():
                                if ig == it_grade and s_f <= mid_size and s_t >= mid_size:
                                    it_tol = tol
                                    break
                        
                        if it_tol is None: continue

                        if tolerance_code == 'js':
                            half_it = it_tol / 2
                            upper_dev = half_it
                            lower_dev = -half_it
                        elif is_es_based:
                            upper_dev = value
                            lower_dev = upper_dev - it_tol
                        elif is_ei_based:
                            lower_dev = value
                            upper_dev = lower_dev + it_tol
                        else:
                            continue # j 或其他
                            
                        key = (size_from, size_to, tolerance_code, it_grade)
                        data[key] = (upper_dev, lower_dev)

        except Exception as e:
            print(f"  讀取 ISO 1 {sheet_name} 失敗: {e}")
            
    return data

def import_shaft_tolerance_iso2(excel_path):
    """匯入軸公差 (ISO 286-2) - 回傳字典"""
    print("  正在讀取 ISO 286-2 軸公差...")
    data = {}
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
    except Exception as e:
        print(f"⚠️ 無法讀取 ISO 286-2 檔案: {e}")
        return {}

    for sheet_name in sheet_names:
        try:
            df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None, engine='openpyxl')
            if len(df) < 4: continue
            
            row0 = df.iloc[0].tolist()
            row1 = df.iloc[1].tolist()
            row2 = df.iloc[2].tolist()
            
            # 取得公差代號 (從 sheet name 或 header)
            current_tol = None
            match = re.search(r'軸([a-z]+?)[至的]', sheet_name)
            if match: current_tol = match.group(1).upper()
            else:
                match = re.search(r'軸([a-z]+)', sheet_name)
                if match: current_tol = match.group(1).upper()
            
            # 識別欄位
            col_map = {}
            
            # 判斷 IT 等級在哪一列 (Row 1 or Row 2)
            it_row = row2
            has_it_row2 = any(extract_it(x) is not None for x in row2[2:])
            if not has_it_row2:
                it_row = row1
            
            for i in range(2, len(row0)):
                val0 = row0[i]
                if pd.notna(val0):
                    s = str(val0).strip()
                    match = re.match(r'^([a-z]+)', s, re.IGNORECASE)
                    if match: current_tol = match.group(1).upper()
                
                if not current_tol: continue
                
                it_num = extract_it(it_row[i])
                if it_num is not None:
                    it_grade = f"IT{it_num}"
                    col_map[i] = (current_tol.lower(), it_grade) # 轉小寫以符合 ISO 1 格式
            
            # 讀取數據
            size_cols = df.columns[:2]
            range_str = df[size_cols[0]].astype(str).fillna('') + ' ' + df[size_cols[1]].astype(str).fillna('')
            size_from_list, size_to_list = zip(*range_str.map(split_range))
            
            num_rows = len(df)
            start_row = 3 if has_it_row2 else 4
            
            for r in range(start_row, num_rows, 2):
                if r + 1 >= num_rows: break
                
                s_from = size_from_list[r]
                s_to = size_to_list[r]
                
                if s_from is None or s_to is None: continue
                
                row_upper = df.iloc[r]
                row_lower = df.iloc[r+1]
                
                for col_idx, (tol_code, it_grade) in col_map.items():
                    val_upper = to_number(row_upper[col_idx])
                    val_lower = to_number(row_lower[col_idx])
                    
                    if val_upper is not None and val_lower is not None:
                        key = (s_from, s_to, tol_code, it_grade)
                        data[key] = (val_upper, val_lower)
                        
        except Exception as e:
            print(f"  處理 ISO 2 {sheet_name} 失敗: {e}")
            
    return data

def import_shaft_tolerance(iso1_path, iso2_path):
    """匯入軸公差 (合併 ISO 1 & 2)"""
    print("\n=== 匯入軸公差 (合併 ISO 1 & 2) ===")
    
    data_iso1 = import_shaft_tolerance_iso1(iso1_path)
    print(f"  ISO 1 數據量: {len(data_iso1)}")
    
    data_iso2 = import_shaft_tolerance_iso2(iso2_path)
    print(f"  ISO 2 數據量: {len(data_iso2)}")
    
    merged_data = data_iso1.copy()
    merged_data.update(data_iso2)
    print(f"  合併後總數據量: {len(merged_data)}")
    
    total_inserted = 0
    with Session() as s:
        with s.begin():
            for (size_from, size_to, code, it), (upper, lower) in merged_data.items():
                ins = mysql_insert(ShaftTolerance).values(
                    size_from_mm=size_from,
                    size_to_mm=size_to,
                    tolerance_code=code,
                    it_grade=it,
                    upper_dev_um=upper,
                    lower_dev_um=lower,
                )
                stmt = ins.on_duplicate_key_update(
                    upper_dev_um=ins.inserted.upper_dev_um,
                    lower_dev_um=ins.inserted.lower_dev_um
                )
                s.execute(stmt)
                total_inserted += 1
                
    print(f"✅ 軸公差匯入完成: {total_inserted} 筆")
    return total_inserted

def clean_invalid_data():
    """清除不合規範的資料 (如 A/B/C > 500mm)"""
    print("\n=== 清除無效資料 ===")
    s = Session()
    try:
        # Delete Shafts a, b, c > 500mm (ISO 286: a,b,c 不適用於 500mm 以上，[500,630] 含端點需保留)
        deleted_shafts = s.query(ShaftTolerance).filter(
            ShaftTolerance.tolerance_code.in_(['a', 'b', 'c']),
            ShaftTolerance.size_from_mm > 500
        ).delete(synchronize_session=False)
        
        # Delete Holes A, B, C > 500mm
        deleted_holes = s.query(HoleTolerance).filter(
            HoleTolerance.tolerance_code.in_(['A', 'B', 'C']),
            HoleTolerance.size_from_mm >= 500
        ).delete(synchronize_session=False)
        
        s.commit()
        print(f"  已刪除軸 a, b, c (> 500mm): {deleted_shafts} 筆")
        print(f"  已刪除孔 A, B, C (> 500mm): {deleted_holes} 筆")
        return deleted_shafts, deleted_holes
        
    except Exception as e:
        s.rollback()
        print(f"❌ 清除失敗: {e}")
        return 0, 0
    finally:
        s.close()

def apply_iso_special_rules():
    """應用 ISO 軸公差特殊規則 (刪除或修正數據)"""
    print("\n=== 應用 ISO 特殊規則 ===")
    s = Session()
    try:
        # Rule 1: Delete a, b where size <= 1mm
        r1 = s.query(ShaftTolerance).filter(
            ShaftTolerance.tolerance_code.in_(['a', 'b']),
            ShaftTolerance.size_to_mm <= 1
        ).delete(synchronize_session=False)
        print(f"  [規則1] 刪除 a, b (<= 1mm): {r1} 筆")

        # Rule 2: Delete h where size <= 1mm AND IT14-16
        # (js 在所有 IT 等級均有定義，不適用此刪除規則)
        r23 = s.query(ShaftTolerance).filter(
            ShaftTolerance.tolerance_code == 'h',
            ShaftTolerance.size_to_mm <= 1,
            ShaftTolerance.it_grade.in_(['IT14', 'IT15', 'IT16'])
        ).delete(synchronize_session=False)
        print(f"  [規則2] 刪除 h (<= 1mm, IT14-16): {r23} 筆")
        
        # Rule 5: Delete v, x, y where size > 500mm (嚴格大於 500，含 500mm 端點需保留)
        r5 = s.query(ShaftTolerance).filter(
            ShaftTolerance.tolerance_code.in_(['v', 'x', 'y']),
            ShaftTolerance.size_from_mm > 500
        ).delete(synchronize_session=False)
        print(f"  [規則5] 刪除 v, x, y (> 500mm): {r5} 筆")

        # Rule 4: Update t (IT5-8, <= 24mm) from u
        sql_r4 = text("""
            UPDATE shaft_tolerance t
            JOIN shaft_tolerance u ON 
                t.size_from_mm = u.size_from_mm AND 
                t.size_to_mm = u.size_to_mm AND 
                t.it_grade = u.it_grade
            SET 
                t.upper_dev_um = u.upper_dev_um,
                t.lower_dev_um = u.lower_dev_um
            WHERE 
                t.tolerance_code = 't' AND 
                u.tolerance_code = 'u' AND
                t.size_to_mm <= 24 AND
                t.it_grade IN ('IT5', 'IT6', 'IT7', 'IT8');
        """)
        res4 = s.execute(sql_r4)
        print(f"  [規則4] 修正 t (<= 24mm, IT5-8) 使用 u 值: {res4.rowcount} 筆")

        # Rule 6: Update v (IT5-8, <= 14mm) from x
        sql_r6 = text("""
            UPDATE shaft_tolerance v
            JOIN shaft_tolerance x ON 
                v.size_from_mm = x.size_from_mm AND 
                v.size_to_mm = x.size_to_mm AND 
                v.it_grade = x.it_grade
            SET 
                v.upper_dev_um = x.upper_dev_um,
                v.lower_dev_um = x.lower_dev_um
            WHERE 
                v.tolerance_code = 'v' AND 
                x.tolerance_code = 'x' AND
                v.size_to_mm <= 14 AND
                v.it_grade IN ('IT5', 'IT6', 'IT7', 'IT8');
        """)
        res6 = s.execute(sql_r6)
        print(f"  [規則6] 修正 v (<= 14mm, IT5-8) 使用 x 值: {res6.rowcount} 筆")

        # Rule 7: Update y (IT6-10, <= 18mm) from z
        sql_r7 = text("""
            UPDATE shaft_tolerance y
            JOIN shaft_tolerance z ON 
                y.size_from_mm = z.size_from_mm AND 
                y.size_to_mm = z.size_to_mm AND 
                y.it_grade = z.it_grade
            SET 
                y.upper_dev_um = z.upper_dev_um,
                y.lower_dev_um = z.lower_dev_um
            WHERE 
                y.tolerance_code = 'y' AND 
                z.tolerance_code = 'z' AND
                y.size_to_mm <= 18 AND
                y.it_grade IN ('IT6', 'IT7', 'IT8', 'IT9', 'IT10');
        """)
        res7 = s.execute(sql_r7)
        print(f"  [規則7] 修正 y (<= 18mm, IT6-10) 使用 z 值: {res7.rowcount} 筆")

        # Rule 8: Delete z, za, zb, zc where size > 500mm (嚴格大於 500，含 500mm 端點需保留)
        r8 = s.query(ShaftTolerance).filter(
            ShaftTolerance.tolerance_code.in_(['z', 'za', 'zb', 'zc']),
            ShaftTolerance.size_from_mm > 500
        ).delete(synchronize_session=False)
        print(f"  [規則8] 刪除 z, za, zb, zc (> 500mm): {r8} 筆")

        s.commit()
        return r1 + r23 + r5 + r8
        
    except Exception as e:
        s.rollback()
        print(f"❌ 規則應用失敗: {e}")
        return 0
    finally:
        s.close()

def apply_iso_hole_rules():
    """應用 ISO 孔公差特殊規則 (刪除或修正數據)"""
    print("\n=== 應用 ISO 孔特殊規則 ===")
    s = Session()
    try:
        # Rule 1: Delete A, B where size <= 1mm
        r1 = s.query(HoleTolerance).filter(
            HoleTolerance.tolerance_code.in_(['A', 'B']),
            HoleTolerance.size_to_mm <= 1
        ).delete(synchronize_session=False)
        print(f"  [孔規則1] 刪除 A, B (<= 1mm): {r1} 筆")

        # Rule 2: Delete H where size <= 1mm AND IT14-18
        r2 = s.query(HoleTolerance).filter(
            HoleTolerance.tolerance_code == 'H',
            HoleTolerance.size_to_mm <= 1,
            HoleTolerance.it_grade.in_(['IT14', 'IT15', 'IT16', 'IT17', 'IT18'])
        ).delete(synchronize_session=False)
        print(f"  [孔規則2] 刪除 H (<= 1mm, IT14-18): {r2} 筆")

        # Rule 3: Delete JS where size <= 1mm AND IT14-16
        r3 = s.query(HoleTolerance).filter(
            HoleTolerance.tolerance_code == 'JS',
            HoleTolerance.size_to_mm <= 1,
            HoleTolerance.it_grade.in_(['IT14', 'IT15', 'IT16'])
        ).delete(synchronize_session=False)
        print(f"  [孔規則3] 刪除 JS (<= 1mm, IT14-16): {r3} 筆")

        # Rule 4: Delete K where size > 3mm AND IT > IT8
        # Note: size_from_mm >= 3 covers ranges starting at 3 (e.g. 3-6), which means >3mm
        it_gt_8 = [f"IT{i}" for i in range(9, 19)]
        r4 = s.query(HoleTolerance).filter(
            HoleTolerance.tolerance_code == 'K',
            HoleTolerance.size_from_mm >= 3,
            HoleTolerance.it_grade.in_(it_gt_8)
        ).delete(synchronize_session=False)
        print(f"  [孔規則4] 刪除 K (> 3mm, > IT8): {r4} 筆")

        # Rule 5: Delete N where size <= 1mm AND IT9-11
        r5 = s.query(HoleTolerance).filter(
            HoleTolerance.tolerance_code == 'N',
            HoleTolerance.size_to_mm <= 1,
            HoleTolerance.it_grade.in_(['IT9', 'IT10', 'IT11'])
        ).delete(synchronize_session=False)
        print(f"  [孔規則5] 刪除 N (<= 1mm, IT9-11): {r5} 筆")

        # Rule 7: Delete V, X, Y where size > 500mm
        r7 = s.query(HoleTolerance).filter(
            HoleTolerance.tolerance_code.in_(['V', 'X', 'Y']),
            HoleTolerance.size_from_mm >= 500
        ).delete(synchronize_session=False)
        print(f"  [孔規則7] 刪除 V, X, Y (> 500mm): {r7} 筆")

        # Rule 10: Delete Z, ZA, ZB, ZC where size > 500mm
        r10 = s.query(HoleTolerance).filter(
            HoleTolerance.tolerance_code.in_(['Z', 'ZA', 'ZB', 'ZC']),
            HoleTolerance.size_from_mm >= 500
        ).delete(synchronize_session=False)
        print(f"  [孔規則10] 刪除 Z, ZA, ZB, ZC (> 500mm): {r10} 筆")

        # Rule 6: Update T (IT5-8, <= 24mm) from U
        sql_r6 = text("""
            UPDATE hole_tolerance t
            JOIN hole_tolerance u ON 
                t.size_from_mm = u.size_from_mm AND 
                t.size_to_mm = u.size_to_mm AND 
                t.it_grade = u.it_grade
            SET 
                t.upper_dev_um = u.upper_dev_um,
                t.lower_dev_um = u.lower_dev_um
            WHERE 
                t.tolerance_code = 'T' AND 
                u.tolerance_code = 'U' AND
                t.size_to_mm <= 24 AND
                t.it_grade IN ('IT5', 'IT6', 'IT7', 'IT8');
        """)
        res6 = s.execute(sql_r6)
        print(f"  [孔規則6] 修正 T (<= 24mm, IT5-8) 使用 U 值: {res6.rowcount} 筆")

        # Rule 8: Update V (IT5-8, <= 14mm) from X
        sql_r8 = text("""
            UPDATE hole_tolerance v
            JOIN hole_tolerance x ON 
                v.size_from_mm = x.size_from_mm AND 
                v.size_to_mm = x.size_to_mm AND 
                v.it_grade = x.it_grade
            SET 
                v.upper_dev_um = x.upper_dev_um,
                v.lower_dev_um = x.lower_dev_um
            WHERE 
                v.tolerance_code = 'V' AND 
                x.tolerance_code = 'X' AND
                v.size_to_mm <= 14 AND
                v.it_grade IN ('IT5', 'IT6', 'IT7', 'IT8');
        """)
        res8 = s.execute(sql_r8)
        print(f"  [孔規則8] 修正 V (<= 14mm, IT5-8) 使用 X 值: {res8.rowcount} 筆")

        # Rule 9: Update Y (IT6-10, <= 18mm) from Z
        sql_r9 = text("""
            UPDATE hole_tolerance y
            JOIN hole_tolerance z ON 
                y.size_from_mm = z.size_from_mm AND 
                y.size_to_mm = z.size_to_mm AND 
                y.it_grade = z.it_grade
            SET 
                y.upper_dev_um = z.upper_dev_um,
                y.lower_dev_um = z.lower_dev_um
            WHERE 
                y.tolerance_code = 'Y' AND 
                z.tolerance_code = 'Z' AND
                y.size_to_mm <= 18 AND
                y.it_grade IN ('IT6', 'IT7', 'IT8', 'IT9', 'IT10');
        """)
        res9 = s.execute(sql_r9)
        print(f"  [孔規則9] 修正 Y (<= 18mm, IT6-10) 使用 Z 值: {res9.rowcount} 筆")

        s.commit()
        return r1 + r2 + r3 + r4 + r5 + r7 + r10
        
    except Exception as e:
        s.rollback()
        print(f"❌ 孔規則應用失敗: {e}")
        return 0
    finally:
        s.close()

if __name__ == "__main__":
    # __file__ 在 server/scripts/，Excel 在 server/data/，所以要往上一層再進 data/
    base_dir = Path(__file__).resolve().parent.parent
    iso1_excel = base_dir / "data" / "ISO_286_1_test.xlsx"
    iso2_excel = base_dir / "data" / "ISO_286_2_test.xlsx"
    iso2_excel_2 = base_dir / "data" / "ISO_286_2_test2.xlsx"
    
    print(f"📁 ISO 1 檔案: {iso1_excel}")
    print(f"📁 ISO 2 檔案: {iso2_excel}")
    print(f"📁 ISO 2 (軸) 檔案: {iso2_excel_2}")
    
    try:
        it_count = import_it_tolerance(iso1_excel)
        hole_count = import_hole_tolerance(iso1_excel, iso2_excel)
        shaft_count = import_shaft_tolerance(iso1_excel, iso2_excel_2)
        
        # 清除無效資料 (A/B/C > 500mm)
        del_shaft, del_hole = clean_invalid_data()
        
        # 應用 ISO 特殊規則 (軸 8條, 孔 10條)
        del_iso_shaft = apply_iso_special_rules()
        del_iso_hole = apply_iso_hole_rules()
        
        # 更新最終計數
        shaft_count = shaft_count - del_shaft - del_iso_shaft
        hole_count = hole_count - del_hole - del_iso_hole
        
        print("\n" + "="*50)
        print("✅ 所有資料匯入與清理完成！")
        print(f"   IT 基本公差: {it_count} 筆")
        print(f"   孔公差: {hole_count} 筆")
        print(f"   軸公差: {shaft_count} 筆")
        print(f"   總計: {it_count + hole_count + shaft_count} 筆")
        print("="*50)
        
    except Exception as e:
        print(f"\n❌ 匯入失敗: {e}")
        import traceback
        traceback.print_exc()

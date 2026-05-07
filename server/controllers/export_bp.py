"""export_bp.py - 報表匯出 Blueprint

包含：
  - POST /api/export_tolerance_csv
  - POST /api/export_contact_lines_csv
  - POST /api/export_analysis_excel
  - POST /api/export_allocation_excel
  - POST /api/run_allocation
  - POST /api/save_allocation
"""

import io
import os
import copy
import pandas as pd
from flask import Blueprint, jsonify, request, Response
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font

export_bp = Blueprint('export', __name__)

# ── Excel 全域樣式常數 ────────────────────────────────────────────────────────
RAD_TO_ARCSEC = 206264.8
PINK_FILL   = PatternFill(start_color='fbcfe8', end_color='fbcfe8', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='fef08a', end_color='fef08a', fill_type='solid')
GREEN_FILL  = PatternFill(start_color='bbf7d0', end_color='bbf7d0', fill_type='solid')
BLUE_FILL   = PatternFill(start_color='bfdbfe', end_color='bfdbfe', fill_type='solid')
RED_FILL    = PatternFill(start_color='fecaca', end_color='fecaca', fill_type='solid')
DBLUE_FILL  = PatternFill(start_color='93c5fd', end_color='93c5fd', fill_type='solid')
FONT_TNR      = Font(name='Times New Roman', size=10)
FONT_TNR_BOLD = Font(name='Times New Roman', size=11, bold=True)


# ── Excel 工具函式 ─────────────────────────────────────────────────────────────

def _write_block(ws, start_row, start_col, title, fill, headers, rows):
    cell = ws.cell(row=start_row, column=start_col)
    cell.value = title
    cell.font  = FONT_TNR_BOLD
    ws.merge_cells(
        start_row=start_row, start_column=start_col,
        end_row=start_row,   end_column=start_col + len(headers) - 1,
    )
    for i, h in enumerate(headers):
        if h:
            c = ws.cell(row=start_row + 1, column=start_col + i)
            c.value = h
            c.font  = FONT_TNR
    for r_idx, r_data in enumerate(rows):
        for c_idx, val in enumerate(r_data):
            c = ws.cell(row=start_row + 2 + r_idx, column=start_col + c_idx)
            c.value = val
            c.font  = FONT_TNR
    last_row = start_row + 2 + len(rows) - 1
    last_col = start_col + len(headers) - 1
    for r in range(start_row, last_row + 1):
        for col in range(start_col, last_col + 1):
            ws.cell(row=r, column=col).fill = fill


def _write_summary_blocks(ws, result: dict):
    """寫入 Tideal、RSS、WC、Sensitivity、Contribution 專業分析塊。"""
    tm = result.get('t_ideal_matrix', [[1,0,0,0],[0,1,0,0],[0,0,1,0],[0,0,0,1]])
    _write_block(ws, 1, 11, 'Tideal Matrix', PINK_FILL, ['','','',''], [
        [tm[0][0], tm[0][1], tm[0][2], tm[0][3]],
        [tm[1][0], tm[1][1], tm[1][2], tm[1][3]],
        [tm[2][0], tm[2][1], tm[2][2], tm[2][3]],
        [tm[3][0], tm[3][1], tm[3][2], tm[3][3]],
    ])
    _write_block(ws, 8,  11, 'Statistics Model(RSS)', YELLOW_FILL, ['tol_range','-3sigma','+3sigma'], [
        ['Xerror', result.get('rss_X', 0)*-1, result.get('rss_X', 0)],
        ['Yerror', result.get('rss_Y', 0)*-1, result.get('rss_Y', 0)],
        ['Zerror', result.get('rss_Z', 0)*-1, result.get('rss_Z', 0)],
    ])
    _write_block(ws, 14, 11, 'Worst Case Model', GREEN_FILL, ['tol_range','min','max'], [
        ['Xerror', result.get('wc_X', 0)*-1, result.get('wc_X', 0)],
        ['Yerror', result.get('wc_Y', 0)*-1, result.get('wc_Y', 0)],
        ['Zerror', result.get('wc_Z', 0)*-1, result.get('wc_Z', 0)],
    ])
    _write_block(ws, 20, 11, 'Angle Statistics Model(arc_second)', YELLOW_FILL, ['tol_range','-3sigma','+3sigma'], [
        ['Xerror', result.get('rss_aX', 0)*-1, result.get('rss_aX', 0)],
        ['Yerror', result.get('rss_aY', 0)*-1, result.get('rss_aY', 0)],
        ['Zerror', result.get('rss_aZ', 0)*-1, result.get('rss_aZ', 0)],
    ])
    _write_block(ws, 26, 11, 'Angle Worst Case Model(arc_second)', GREEN_FILL, ['tol_range','min','max'], [
        ['Xerror', result.get('wc_aX', 0)*-1, result.get('wc_aX', 0)],
        ['Yerror', result.get('wc_aY', 0)*-1, result.get('wc_aY', 0)],
        ['Zerror', result.get('wc_aZ', 0)*-1, result.get('wc_aZ', 0)],
    ])
    def _rows(data_list):
        return [[d.get('name',''), d.get('x',0), d.get('y',0), d.get('z',0)] for d in data_list]
    _write_block(ws, 1, 16, 'Sensitivity Analysis',       BLUE_FILL,  ['tol_sym','X(%)','Y(%)','Z(%)'], _rows(result.get('sensitivity',        [])))
    _write_block(ws, 1, 22, 'Contribution Analysis',      RED_FILL,   ['tol_sym','X(%)','Y(%)','Z(%)'], _rows(result.get('contribution',       [])))
    _write_block(ws, 1, 28, 'Angle Sensitivity Analysis', DBLUE_FILL, ['tol_sym','X(%)','Y(%)','Z(%)'], _rows(result.get('angle_sensitivity',  [])))
    _write_block(ws, 1, 34, 'Angle Contribution Analysis',DBLUE_FILL, ['tol_sym','X(%)','Y(%)','Z(%)'], _rows(result.get('angle_contribution', [])))


# ── 路由 ──────────────────────────────────────────────────────────────────────

@export_bp.post('/api/export_tolerance_csv')
def export_tolerance_csv():
    data      = request.get_json()
    path_data = data.get('pathData', [])
    lang      = data.get('lang', 'zh-TW')
    en        = (lang == 'en')

    rows = []
    for item in path_data:
        code = item.get('name') if item.get('type') == 'feature' else item.get('axis')
        rows.append({
            ('Path Code'               if en else '路徑代碼'):                code,
            ('Value(tra/rot/tol)'      if en else '數值(平移、旋轉、公差值)'): item.get('val', 0.01),
            ('Bias(offset)'            if en else '偏差值(公差帶偏移值)'):     item.get('bias', 0),
            ('Ang Tol Dist'            if en else '角度公差轉換距離'):         item.get('dist', '') or '',
            ('Nominal Size'            if en else '公稱尺寸'):                item.get('nominal_size', '') or '',
            ('IT Grade'                if en else 'IT等級'):                  item.get('it_grade', '') or '',
            ('Part Name'               if en else '所屬零件'):                item.get('part', '') or '',
            ('Tol Type'                if en else '公差類型'):                item.get('tol_type', '') or '',
        })

    df  = pd.DataFrame(rows)
    buf = io.StringIO()
    df.to_csv(buf, index=False, encoding='utf-8-sig')
    return Response(
        buf.getvalue(),
        mimetype='text/csv',
        headers={'Content-disposition': 'attachment; filename=Tolerance_Path_Export.csv'},
    )


@export_bp.post('/api/export_contact_lines_csv')
def export_contact_lines_csv():
    data  = request.get_json()
    pairs = data.get('pairs', [])
    lang  = data.get('lang', 'zh-TW')
    en    = (lang == 'en')

    rows = [{
        ('Feature 1'       if en else '特徵面 1'): pair.get('start'),
        ('Feature 2'       if en else '特徵面 2'): pair.get('end'),
        ('Connection Type' if en else '連結類型'): 'Hard Contact' if en else '硬接觸',
    } for pair in pairs]

    df  = pd.DataFrame(rows)
    buf = io.StringIO()
    df.to_csv(buf, index=False, encoding='utf-8-sig')
    return Response(
        buf.getvalue(),
        mimetype='text/csv',
        headers={'Content-disposition': 'attachment; filename=Contact_Lines_Export.csv'},
    )


@export_bp.post('/api/export_analysis_excel')
def export_analysis_excel():
    body      = request.get_json(force=True)
    path_data = body.get('pathData', [])
    result    = body.get('result', {})
    lang      = body.get('lang', 'zh-TW')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Tolerance_Report'

    hdr = ['Tolerance Code','Value','Bias','Dist'] if lang == 'en' else ['路徑代碼','數值','偏差','轉換距離']
    for col, h in enumerate(hdr, 1):
        c = ws.cell(row=1, column=col)
        c.value = h
        c.font  = FONT_TNR_BOLD

    for r_idx, item in enumerate(path_data, 2):
        code = item.get('name') if item.get('type') == 'feature' else item.get('axis', '')
        ws.cell(row=r_idx, column=1).value = code
        ws.cell(row=r_idx, column=2).value = item.get('val', 0)
        ws.cell(row=r_idx, column=3).value = item.get('bias', 0)
        ws.cell(row=r_idx, column=4).value = item.get('dist', 1)

    _write_summary_blocks(ws, result)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=Tolerance_Analysis_Report.xlsx'},
    )


@export_bp.post('/api/export_allocation_excel')
def export_allocation_excel():
    body       = request.get_json(force=True)
    prev_path  = body.get('prevPathData', [])
    new_path   = body.get('newPathData', [])
    report     = body.get('report', {})
    new_result = body.get('analysisResult', {})
    lang       = body.get('lang', 'zh-TW')
    en         = (lang == 'en')

    wb  = Workbook()
    ws1 = wb.active
    ws1.title = 'Allocation_Comparison'

    ws1.cell(row=1, column=1).value = '[RSS Improvement Summary]' if en else '【RSS 預測改善匯總】'
    ws1.cell(row=1, column=1).font  = FONT_TNR_BOLD

    hdr_rss = (['Axis','Before','After','Improve %'] if en
               else ['軸向','目前 (Before)','優化後 (After)','改善 (%)'])
    for i, h in enumerate(hdr_rss, 1):
        ws1.cell(row=2, column=i).value = h
        ws1.cell(row=2, column=i).fill  = YELLOW_FILL

    curr_r = 3
    for ax in ['X','Y','Z','aX','aY','aZ']:
        item = report.get(ax)
        if not item: continue
        ws1.cell(row=curr_r, column=1).value = ax
        ws1.cell(row=curr_r, column=2).value = item.get('rss_before', 0)
        ws1.cell(row=curr_r, column=3).value = item.get('rss_after', 0)
        ws1.cell(row=curr_r, column=4).value = f"{item.get('rss_improve_pct', 0)}%"
        curr_r += 1

    curr_r += 2
    ws1.cell(row=curr_r, column=1).value = '[Tolerance Detail Table]' if en else '【個別公差明細表】'
    ws1.cell(row=curr_r, column=1).font  = FONT_TNR_BOLD
    curr_r += 1

    hdr_tol = (['Item','Old Tol','New Tol','Delta %','Quadrant'] if en
               else ['項目','原公差','新公差','變動 (%)','診斷象限'])
    for i, h in enumerate(hdr_tol, 1):
        ws1.cell(row=curr_r, column=i).value = h
        ws1.cell(row=curr_r, column=i).fill  = BLUE_FILL

    prev_map = {p.get('name'): p.get('val', 0) for p in prev_path if p.get('name')}
    q_names  = ({1:'Q1',2:'Q2',3:'Q3',4:'Q4'} if en
                else {1:'Q1(關鍵)',2:'Q2(維護)',3:'Q3(次要)',4:'Q4(放寬)'})

    for item in new_path:
        if item.get('type') != 'feature': continue
        curr_r += 1
        name    = item.get('name')
        new_val = item.get('val', 0)
        old_val = prev_map.get(name, new_val)
        delta   = ((new_val - old_val) / old_val * 100) if old_val > 0 else 0
        q       = item.get('quadrant', 4)
        ws1.cell(row=curr_r, column=1).value = name
        ws1.cell(row=curr_r, column=2).value = old_val
        ws1.cell(row=curr_r, column=3).value = new_val
        ws1.cell(row=curr_r, column=4).value = f'{round(delta, 1)}%'
        ws1.cell(row=curr_r, column=5).value = q_names.get(q, 'Q4')

    ws2 = wb.create_sheet('Optimized_Analysis')
    for r_idx, item in enumerate(new_path, 2):
        code = item.get('name') if item.get('type') == 'feature' else item.get('axis', '')
        ws2.cell(row=r_idx, column=1).value = code
        ws2.cell(row=r_idx, column=2).value = item.get('val', 0)
        ws2.cell(row=r_idx, column=3).value = item.get('bias', 0)
        ws2.cell(row=r_idx, column=4).value = item.get('dist', 1)
    _write_summary_blocks(ws2, new_result)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=Tolerance_Allocation_Report.xlsx'},
    )


@export_bp.post('/api/run_allocation')
def run_allocation():
    body            = request.get_json(force=True)
    mode            = body.get('mode', 'auto')
    path_data       = body.get('pathData', [])
    analysis_result = body.get('analysisResult', {})
    baseline        = body.get('baseline', {})

    try:
        from analysis_service import compute_allocation, compare_allocation
        import analysis_service

        if mode == 'auto':
            target   = float(body.get('target', 0.05))
            strategy = body.get('weight', 'medium')
            axis     = body.get('axis', 'Z')

            # 重新分析當前路徑，取得調配前的真實基準值
            # （路徑可能在上次分析後被 IT 調整或配合建議修改過）
            before_res = analysis_service.analyze_tolerance_path(path_data)

            res_data = compute_allocation(path_data, before_res, target, strategy, axis)

            overrides = {}
            quadrants = {}
            q_map     = res_data.get('report', {}).get('quadrants', {})
            for item in res_data['newPathData']:
                if item.get('type') == 'feature':
                    name = item.get('name')
                    overrides[name] = item.get('val', 0)
                    if name in q_map:
                        quadrants[name] = q_map[name]

            from scripts.dsl_builder import build_full_dsl
            updated_dsl = build_full_dsl(mode='network', tolerance_overrides=overrides, quadrants=quadrants)

            final_res  = analysis_service.analyze_tolerance_path(res_data['newPathData'])

            def _ipct(b, a):
                return 0.0 if b == 0 else round((b - a) / b * 100, 2)

            full_report = {}
            for ax in ['X','Y','Z','aX','aY','aZ']:
                rb  = before_res.get(f'rss_{ax}', 0) or 0   # 當前路徑真實前值
                ra  = final_res.get(f'rss_{ax}',  0) or 0
                wb2 = before_res.get(f'wc_{ax}',  0) or 0   # 當前路徑真實前值
                wa  = final_res.get(f'wc_{ax}',   0) or 0
                full_report[ax] = {
                    'rss_before': round(rb, 6), 'rss_after': round(ra, 6),
                    'rss_improve_pct': _ipct(rb, ra),
                    'wc_before': round(wb2, 6), 'wc_after': round(wa, 6),
                    'wc_improve_pct':  _ipct(wb2, wa),
                }
            full_report['quadrants'] = q_map

            return jsonify({
                'ok': True, 'mode': 'auto',
                'newPathData': res_data['newPathData'],
                'report': full_report, 'dsl': updated_dsl,
                'analysisResult': final_res,
                'beforeResult': before_res,   # 回傳真實前值供前端更新
            })

        elif mode == 'compare':
            current_analysis = analysis_service.analyze_tolerance_path(path_data)
            res_data         = compare_allocation(baseline, current_analysis)
            names  = [i['name'] for i in path_data if i.get('type') == 'feature']
            q_map, _ = analysis_service.get_quadrant_analysis(current_analysis, names, body.get('axis', 'Z'))
            res_data['quadrants'] = q_map
            updated_path = copy.deepcopy(path_data)
            for item in updated_path:
                if item.get('type') == 'feature' and item.get('name') in q_map:
                    item['quadrant'] = q_map[item['name']]
            return jsonify({
                'ok': True, 'mode': 'compare',
                'report': res_data, 'newPathData': updated_path,
                'analysisResult': current_analysis,
            })

        return jsonify({'ok': False, 'error': f'未知模式: {mode}'}), 400

    except Exception as e:
        import traceback
        return jsonify({'ok': False, 'error': str(e), 'trace': traceback.format_exc()}), 500


@export_bp.post('/api/save_allocation')
def save_allocation():
    body      = request.get_json(force=True)
    path_data = body.get('pathData', [])
    result    = body.get('result', {})
    lang      = body.get('lang', 'zh-TW')
    excel_path = 'test.xlsx'

    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
    else:
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

    alloc_sheets = [s for s in wb.sheetnames if s.startswith('Allocation_V')]
    sheet_name   = f'Allocation_V{len(alloc_sheets) + 1}'
    ws = wb.create_sheet(sheet_name)

    hdr = ['Path Code','Value','Bias','Dist'] if lang == 'en' else ['路徑代碼','數值','偏差','轉換距離']
    for col, h in enumerate(hdr, 1):
        ws.cell(row=1, column=col).value = h

    for i, item in enumerate(path_data, 2):
        code = item.get('name') if item.get('type') == 'feature' else item.get('axis', '')
        ws.cell(row=i, column=1).value = code
        ws.cell(row=i, column=2).value = item.get('val', 0)
        ws.cell(row=i, column=3).value = item.get('bias', 0)
        ws.cell(row=i, column=4).value = item.get('dist', 1)

    ws.cell(row=1, column=6).value = 'Analysis Results'
    ws.cell(row=2, column=6).value = 'RSS (±3σ)'
    ws.cell(row=2, column=7).value = result.get('rss_X_std', 0) * 3

    try:
        wb.save(excel_path)
        return jsonify({'ok': True, 'msg': f'Saved to {sheet_name}'})
    except Exception as e:
        return jsonify({'ok': False, 'msg': str(e)}), 500
